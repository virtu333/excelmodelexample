import openpyxl

wb = openpyxl.load_workbook("arctura_5partner_gtm_model_v1.xlsx", data_only=True)

# 1. Check Non-GAAP Impl+Consulting total on Anchor Revenue sheet
ar = wb["Anchor Revenue"]

# Find the "TOTAL NON-GAAP IMPL + CONSULTING" row
print("=== NON-GAAP IMPL + CONSULTING SECTION ===")
for row in ar.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and "NON-GAAP IMPL" in str(cell.value):
        r = cell.row
        print(f"Row {r}: {cell.value}")
        # Print all quarterly values (cols C through V = 3 through 22)
        vals = []
        for c in range(3, 23):
            v = ar.cell(row=r, column=c).value
            vals.append(v if v else 0)
        print(f"  Quarterly values: {vals}")
        total = sum(v for v in vals if isinstance(v, (int, float)))
        print(f"  Total: ${total:,.0f}")

# 2. Check annual summary - find Non-GAAP Impl + Consulting annual row
print("\n=== ANNUAL SUMMARY - NON-GAAP IMPL + CONSULTING ===")
for row in ar.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and "Non-GAAP Impl + Consulting" in str(cell.value) and cell.row > 900:
        r = cell.row
        print(f"Row {r}: {cell.value}")
        for c in range(3, 8):  # C through G = years 2027-2031
            v = ar.cell(row=r, column=c).value
            yr = 2027 + (c - 3)
            print(f"  {yr}: ${v:,.0f}" if v else f"  {yr}: $0")

# 3. Check that GAAP totals haven't changed (subscription unchanged)
print("\n=== ANNUAL GAAP TOTAL ===")
for row in ar.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and cell.value == "TOTAL ANNUAL REVENUE":
        r = cell.row
        for c in range(3, 8):
            v = ar.cell(row=r, column=c).value
            yr = 2027 + (c - 3)
            print(f"  {yr}: ${v:,.0f}" if v else f"  {yr}: $0")

print("\n=== ANNUAL TOTAL incl. Non-GAAP ===")
for row in ar.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and cell.value == "TOTAL incl. Non-GAAP":
        r = cell.row
        for c in range(3, 8):
            v = ar.cell(row=r, column=c).value
            yr = 2027 + (c - 3)
            print(f"  {yr}: ${v:,.0f}" if v else f"  {yr}: $0")

# 4. Check Combined Summary Non-GAAP row
print("\n=== COMBINED SUMMARY - NON-GAAP ROW (row 6) ===")
cs = wb["Combined Summary"]
for c in range(3, 8):
    v = cs.cell(row=6, column=c).value
    yr = 2027 + (c - 3)
    print(f"  {yr}: ${v:,.0f}" if v else f"  {yr}: $0")

# 5. Check Valuation - Anchor
print("\n=== VALUATION - ANCHOR ===")
vl = wb["Valuation - Anchor"]
for row in vl.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and ("Enterprise" in str(cell.value) or "Non-GAAP Impl" in str(cell.value) or "Total Anchor" in str(cell.value)):
        r = cell.row
        v = vl.cell(row=r, column=3).value
        print(f"Row {r}: {cell.value} = {v}")

wb.close()
