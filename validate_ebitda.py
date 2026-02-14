import openpyxl

wb = openpyxl.load_workbook("arctura_5partner_gtm_model_v1.xlsx", data_only=True)

# 1. Check Assumptions new inputs
print("=== ASSUMPTIONS ===")
assume = wb["Assumptions"]
for row, label in [(129, "G&A %"), (130, "R&D %"), (131, "Amort Years"), (132, "EBITDA Mult")]:
    v = assume.cell(row=row, column=3).value
    print(f"  Row {row} ({label}): {v}")

# 2. Check P&L Summary
print("\n=== P&L SUMMARY ===")
pnl = wb["P&L Summary"]
rows_to_check = [
    (10, "TOTAL REVENUE (GAAP)"),
    (17, "TOTAL COGS"),
    (20, "Gross Profit"),
    (21, "Gross Margin %"),
    (24, "Anchor NRC Amort"),
    (25, "Expansion NRC Amort"),
    (26, "Total Depreciation"),
    (29, "G&A OpEx"),
    (30, "R&D OpEx"),
    (31, "Total OpEx"),
    (34, "EBITDA"),
    (35, "EBITDA Margin %"),
]
for row, label in rows_to_check:
    vals = []
    for c in range(3, 8):
        v = pnl.cell(row=row, column=c).value
        vals.append(v)
    yr_labels = [2027, 2028, 2029, 2030, 2031]
    formatted = []
    for yr, v in zip(yr_labels, vals):
        if v is None:
            formatted.append(f"{yr}: -")
        elif isinstance(v, float) and abs(v) < 1:
            formatted.append(f"{yr}: {v:.1%}")
        else:
            formatted.append(f"{yr}: ${v:,.0f}" if isinstance(v, (int, float)) else f"{yr}: {v}")
    print(f"  Row {row} ({label}): {' | '.join(formatted)}")

# 3. Check Valuation - Anchor
print("\n=== VALUATION - ANCHOR ===")
vl = wb["Valuation - Anchor"]
for row in vl.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and any(kw in str(cell.value) for kw in ["Enterprise Value", "EBITDA Margin", "Gross Margin", "COMPARISON"]):
        r = cell.row
        v = vl.cell(row=r, column=3).value
        if v is not None:
            if isinstance(v, float) and abs(v) < 1:
                print(f"  Row {r}: {cell.value} = {v:.1%}")
            else:
                print(f"  Row {r}: {cell.value} = ${v:,.0f}" if isinstance(v, (int, float)) else f"  Row {r}: {cell.value} = {v}")
        else:
            print(f"  Row {r}: {cell.value} = (no value in C)")

# 4. Check Valuation - Total
print("\n=== VALUATION - TOTAL ===")
vt = wb["Valuation - Total"]
for row in vt.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and any(kw in str(cell.value) for kw in ["Enterprise Value", "EBITDA Margin", "Gross Margin", "COMPARISON"]):
        r = cell.row
        v = vt.cell(row=r, column=3).value
        if v is not None:
            if isinstance(v, float) and abs(v) < 1:
                print(f"  Row {r}: {cell.value} = {v:.1%}")
            else:
                print(f"  Row {r}: {cell.value} = ${v:,.0f}" if isinstance(v, (int, float)) else f"  Row {r}: {cell.value} = {v}")
        else:
            print(f"  Row {r}: {cell.value} = (no value in C)")

# 5. Revenue-based EV check
print("\n=== REVENUE EV CHECK ===")
for row in vl.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and "Revenue" in str(cell.value) and "EV" in str(cell.value):
        v = vl.cell(row=cell.row, column=3).value
        print(f"  Anchor Rev EV: ${v:,.0f}" if v else f"  Anchor Rev EV: -")
for row in vt.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and "Revenue" in str(cell.value) and "EV" in str(cell.value):
        v = vt.cell(row=cell.row, column=3).value
        print(f"  Total Rev EV: ${v:,.0f}" if v else f"  Total Rev EV: -")

wb.close()
