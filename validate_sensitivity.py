import openpyxl

wb = openpyxl.load_workbook("arctura_5partner_gtm_model_v1.xlsx", data_only=True)

def find_rows(ws, labels):
    """Find rows by partial label match in column B."""
    results = {}
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if cell.value:
            for label in labels:
                if label.lower() in str(cell.value).lower() and label not in results:
                    results[label] = cell.row
    return results

def get_annual(ws, row):
    """Get 5 annual values from columns C-G."""
    vals = []
    for c in range(3, 8):
        v = ws.cell(row=row, column=c).value
        vals.append(v if isinstance(v, (int, float)) else 0)
    return vals

def fmt(v):
    if v is None or v == 0:
        return "-"
    return f"${v:,.0f}"

# === ANCHOR REVENUE ===
ar = wb["Anchor Revenue"]
print("=" * 90)
print("ANCHOR REVENUE (Annual Summary)")
print("=" * 90)

# Search in the annual summary area (rows > 900 typically)
anchor_labels = {}
for row in ar.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and cell.row > 100:
        val = str(cell.value).strip()
        if val.startswith("Implementation") and "Implementation" not in anchor_labels:
            anchor_labels["Implementation"] = cell.row
        elif "Non-GAAP Impl" in val and "Impl+Consult" not in anchor_labels:
            anchor_labels["Impl+Consult"] = cell.row
        elif "Non-GAAP Sub" in val and "Subscription" not in anchor_labels:
            anchor_labels["Subscription"] = cell.row
        elif val.startswith("Licensing") and "Licensing" not in anchor_labels:
            anchor_labels["Licensing"] = cell.row
        elif val.startswith("Operations") and "Operations" not in anchor_labels:
            anchor_labels["Operations"] = cell.row
        elif val.startswith("Maintenance") and "Maintenance" not in anchor_labels:
            anchor_labels["Maintenance"] = cell.row
        elif val.startswith("Inspection") and "Inspection" not in anchor_labels:
            anchor_labels["Inspection"] = cell.row
        elif val == "TOTAL ANNUAL REVENUE":
            anchor_labels["GAAP Total"] = cell.row
        elif "incl. Non-GAAP" in val and "TOTAL" in val:
            anchor_labels["incl NonGAAP"] = cell.row

years = [2027, 2028, 2029, 2030, 2031]
print(f"{'Stream':<25} {'2027':>12} {'2028':>12} {'2029':>12} {'2030':>12} {'2031':>12} {'5yr Total':>14}")
print("-" * 90)

for label, row in anchor_labels.items():
    vals = get_annual(ar, row)
    total = sum(vals)
    print(f"{label:<25} {fmt(vals[0]):>12} {fmt(vals[1]):>12} {fmt(vals[2]):>12} {fmt(vals[3]):>12} {fmt(vals[4]):>12} {fmt(total):>14}")

# === EXPANSION PIPELINE ===
exp = wb["Expansion Pipeline"]
print("\n" + "=" * 90)
print("EXPANSION PIPELINE (Annual Summary)")
print("=" * 90)

panchor_labels = {}
for row in exp.iter_rows(min_col=2, max_col=2, values_only=False):
    cell = row[0]
    if cell.value and cell.row > 30:
        val = str(cell.value).strip()
        if "Consult" in val and "Impl" in val and "C+I" not in panchor_labels:
            panchor_labels["C+I"] = cell.row
        elif "Subscription" in val and "Sub" not in panchor_labels:
            panchor_labels["Sub"] = cell.row
        elif "Utilization" in val and "TOTAL" not in val and "Util" not in panchor_labels:
            panchor_labels["Util"] = cell.row
        elif val == "TOTAL EXPANSION ANNUAL":
            panchor_labels["Expansion Total"] = cell.row

print(f"{'Stream':<25} {'2027':>12} {'2028':>12} {'2029':>12} {'2030':>12} {'2031':>12} {'5yr Total':>14}")
print("-" * 90)

for label, row in panchor_labels.items():
    vals = get_annual(exp, row)
    total = sum(vals)
    print(f"{label:<25} {fmt(vals[0]):>12} {fmt(vals[1]):>12} {fmt(vals[2]):>12} {fmt(vals[3]):>12} {fmt(vals[4]):>12} {fmt(total):>14}")

# === COMBINED ===
cs = wb["Combined Summary"]
print("\n" + "=" * 90)
print("COMBINED TOTALS")
print("=" * 90)
for r in [5, 6, 7, 10, 11]:
    label = cs.cell(row=r, column=2).value or ""
    vals = get_annual(cs, r)
    total = sum(vals)
    print(f"{label.strip():<35} {fmt(vals[0]):>12} {fmt(vals[1]):>12} {fmt(vals[2]):>12} {fmt(vals[3]):>12} {fmt(vals[4]):>12} {fmt(total):>14}")

# === VALUATION ===
print("\n" + "=" * 90)
print("ENTERPRISE VALUES")
print("=" * 90)
for sheet_name in ["Valuation - Anchor", "Valuation - Total"]:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=False):
        cell = row[0]
        if cell.value and "Enterprise Value" in str(cell.value):
            v = ws.cell(row=cell.row, column=3).value
            print(f"  {sheet_name} | {cell.value}: ${v:,.0f}" if v else f"  {sheet_name} | {cell.value}: -")

# === P&L Key Lines ===
print("\n" + "=" * 90)
print("P&L KEY METRICS")
print("=" * 90)
pnl = wb["P&L Summary"]
pnl_rows = [(10, "Revenue"), (17, "COGS"), (20, "Gross Profit"), (21, "Gross Margin %"),
            (26, "Depreciation"), (31, "OpEx"), (34, "EBITDA"), (35, "EBITDA Margin %")]
print(f"{'Metric':<25} {'2027':>12} {'2028':>12} {'2029':>12} {'2030':>12} {'2031':>12}")
print("-" * 85)
for row, label in pnl_rows:
    vals = get_annual(pnl, row)
    formatted = []
    for v in vals:
        if label.endswith("%"):
            formatted.append(f"{v:.1%}" if isinstance(v, (int, float)) and v != 0 else "-")
        else:
            formatted.append(fmt(v))
    print(f"{label:<25} {formatted[0]:>12} {formatted[1]:>12} {formatted[2]:>12} {formatted[3]:>12} {formatted[4]:>12}")

wb.close()
