"""
Windows-compatible Excel formula recalculation using LibreOffice.
Recalculates all formulas, saves, then scans for errors.
"""
import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
MACRO_DIR = os.path.join(os.environ.get("APPDATA", ""), "LibreOffice", "4", "user", "basic", "Standard")
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def setup_macro():
    macro_file = os.path.join(MACRO_DIR, MACRO_FILENAME)
    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True
    if not os.path.exists(MACRO_DIR):
        subprocess.run([SOFFICE, "--headless", "--terminate_after_init"],
                       capture_output=True, timeout=15)
        os.makedirs(MACRO_DIR, exist_ok=True)
    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception as e:
        print(f"Failed to install macro: {e}")
        return False


def scan_errors(filename):
    wb = load_workbook(filename, data_only=True)
    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {err: [] for err in excel_errors}
    total_errors = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in excel_errors:
                        if err in cell.value:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    wb2 = load_workbook(filename, data_only=False)
    formula_count = 0
    for sheet_name in wb2.sheetnames:
        ws = wb2[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    wb2.close()

    result = {"status": "success" if total_errors == 0 else "errors_found",
              "total_errors": total_errors, "total_formulas": formula_count, "error_summary": {}}
    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {"count": len(locations), "locations": locations[:20]}
    return result


def recalc(filename, timeout=60):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}
    abs_path = str(Path(filename).absolute())

    if not setup_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [SOFFICE, "--headless", "--norestore",
           "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
           abs_path]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        if result.returncode != 0:
            return {"error": result.stderr or "Unknown error during recalculation"}
    except subprocess.TimeoutExpired:
        return {"error": f"Recalculation timed out after {timeout}s"}

    return scan_errors(filename)


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc_win.py <excel_file> [timeout_seconds]")
        sys.exit(1)
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
