"""
Arctura Revenue Model v1 — Build Script
Creates arctura_5partner_gtm_model_v1.xlsx from scratch using openpyxl.
"""
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Constants ---
OUTPUT_FILE = "arctura_5partner_gtm_model_v1.xlsx"
QUARTERS = [f"Q{q}'{y}" for y in range(27, 32) for q in range(1, 5)]  # Q1'27..Q4'31
NUM_Q = 20
FONT_NAME = "Arial"

# Number formats
CUR_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%'
INT_FMT = '0'
MULT_FMT = '0.0"x"'

# Colors
BLUE_FONT = Font(name=FONT_NAME, color="0000FF")
BLACK_FONT = Font(name=FONT_NAME, color="000000")
# GREEN_FONT removed — use BLACK_FONT instead (no green row labels)
WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", bold=True)
BOLD_FONT = Font(name=FONT_NAME, bold=True)
BOLD_BLUE = Font(name=FONT_NAME, color="0000FF", bold=True)
# BOLD_GREEN removed — use BOLD_FONT instead (no green row labels)
GRAY_SMALL = Font(name=FONT_NAME, color="808080", size=9)
GRAY_BOLD = Font(name=FONT_NAME, color="808080", bold=True)
DARK_BLUE_FILL = PatternFill("solid", fgColor="003366")
YELLOW_FILL = PatternFill("solid", fgColor="F9E2A1")
LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")

# --- Assumptions Sheet Row Map ---
# Rows 1-27 are unchanged from original layout
A_ZONES = 10           # Number of Zones
A_PLAN_Q = 11         # Planning Start Quarter
A_GOLIVE_Q = 12       # Go-Live Quarter
A_LICENSE_Q = 13      # License Activation Quarter
A_IMPL_MSRP = 14      # Implementation MSRP (formula)
A_LIC_FEE = 15        # Annual Licensing Fee (formula)
A_PERZONE_MSRP = 17    # Per-Zone Implementation MSRP
A_PERZONE_LIC = 18     # Per-Zone Annual License
A_SIGNING_PCT = 24    # % at Signing
A_GOLIVE_PCT = 25     # % at Go-Live
A_MONTH12_PCT = 26    # % at Month 12

# Zone Go-Live Ramp grid (rows 28-37, inserted)
A_ZONE_RAMP_HDR = 28
A_ZONE_RAMP_QHDR = 29
A_ZONE_RAMP_START = 30  # Partner 1 ramp row; partners 2-5 at 31-34
A_ZONE_RAMP_TOTAL = 35
A_ZONE_RAMP_CUM = 36

# Rows below shifted +10 from original
A_QI_LOOKUP_HDR = 38
A_PLAN_IDX = 39       # was 29
A_GOLIVE_IDX = 40     # was 30
A_LICENSE_IDX = 41    # was 31

A_EXP_HDR = 43
A_EXP_IMPL_FEE = 44   # was 34
A_EXP_CONSULT = 45     # was 35
A_EXP_SUB = 46         # was 36
A_EXP_IMPL_MO = 47     # was 37
A_EXP_IMPL_Q = 48      # was 38

A_PIPE_HDR = 50
A_PIPE_QHDR = 51       # was 41
A_PIPE_NEWSYS = 52      # was 42
A_PIPE_AVGZONES = 53    # was 43
A_PIPE_NEWZONES = 54    # was 44
A_PIPE_CUMZONES = 55    # was 45
A_PIPE_GOLIVE = 56      # was 46
A_PIPE_CUMGOLIVE = 57   # was 47

A_UTIL_HDR = 59
A_TIER1 = 60            # was 50
A_TIER2 = 61            # was 51
A_MAINT_PRICE = 62        # was 52
A_INSP_PRICE = 63        # was 53

A_UC_HDR = 65
A_UC_COLS = 66          # column headers row
A_UC_START = 67         # was 57 - Autonomous Picking

A_RAMP_HDR = 76
A_RAMP_START = 77       # was 67 - Q1 ramp value
A_RAMP_END = 84         # was 74 - Q8 ramp value
A_RAMP_RANGE = "$C$77:$C$84"

A_DECAY_HDR = 86
A_DECAY_ZONE_START = 87  # was 77
A_DECAY_6 = 93           # 6-zone blended avg
A_DECAY_5 = 94           # 5-zone blended avg
A_DECAY_4 = 95           # 4-zone blended avg
A_DECAY_3 = 96           # 3-zone blended avg
A_DECAY_2 = 97           # 2-zone blended avg

A_MAINT_HDR = 99
A_ROBOTS = 100             # was 87
A_MAINT_RAMP_START = 101   # was 88
A_MAINT_RAMP_Q = 102       # was 89

A_INSP_HDR = 104
A_SCANS = 105          # was 92

A_SCENARIO_HDR = 107
A_SCENARIO_COLS = 108    # was 95
A_SCENARIO_START = 109   # was 96

A_COST_HDR = 116         # was 103
A_NRC_FIRST = 117        # was 104
A_MRC = 118              # was 105
A_DISC = 119             # was 106
A_ARC = 120              # was 107
A_QRC = 121              # was 108
A_NRC_ADDL = 122         # was 109
A_QRC_ADDL = 123         # was 110
A_BLEND_NRC = 125        # was 112
A_BLEND_QRC = 126        # was 113

# New: Operating Expense & Valuation Assumptions
A_OPEX_HDR = 128
A_GA_PCT = 129           # G&A OpEx % of Revenue: 10%
A_RD_PCT = 130           # R&D OpEx % of Revenue: 20%
A_AMORT_YRS = 131        # CapEx Amortization Period: 5 years
A_EBITDA_MULT = 132      # EBITDA Terminal Multiple: 15x

A_SOURCES_NOTE = 134     # was 125, shifted +6
A_SOURCES_START = 135    # was 126, shifted +6

# System data (license dates per v0.4)
SYSTEMS = [
    {"name": "Partner 1 Atlas",     "col": "C", "zones": 6, "plan": "Q2'27", "golive": "Q4'27", "license": "Q2'29"},
    {"name": "Partner 2 Titan",     "col": "D", "zones": 4, "plan": "Q3'27", "golive": "Q1'28", "license": "Q1'29"},
    {"name": "Partner 3 Apex",      "col": "E", "zones": 5, "plan": "Q4'27", "golive": "Q2'28", "license": "Q2'29"},
    {"name": "Partner 4 Vanguard",  "col": "F", "zones": 3, "plan": "Q1'28", "golive": "Q3'28", "license": "Q3'29"},
    {"name": "Partner 5 Nimble",    "col": "G", "zones": 2, "plan": "Q2'28", "golive": "Q4'28", "license": "Q4'29"},
]

# Zone go-live ramp defaults (Q1'27 through Q4'28 = 8 quarters)
ZONE_RAMP_DATA = [
    [0, 0, 0, 3, 3, 0, 0, 0],  # Partner 1 Atlas (6 total)
    [0, 0, 0, 0, 2, 2, 0, 0],  # Partner 2 Titan (4 total)
    [0, 0, 0, 0, 0, 2, 3, 0],  # Partner 3 Apex (5 total)
    [0, 0, 0, 0, 0, 0, 2, 1],  # Partner 4 Vanguard (3 total)
    [0, 0, 0, 0, 0, 0, 0, 2],  # Partner 5 Nimble (2 total)
]

# Use cases (row references use shifted Assumptions rows)
USE_CASES = [
    {"name": "Autonomous Picking",         "row": A_UC_START,     "tier": "Tier 1", "qtrs_after": 0, "avg_min": 25,  "ops": 18.0},
    {"name": "Pallet Sorting",             "row": A_UC_START + 1, "tier": "Tier 2", "qtrs_after": 1, "avg_min": 15,  "ops": 24.0},
    {"name": "Inventory Scanning",         "row": A_UC_START + 2, "tier": "Tier 1", "qtrs_after": 2, "avg_min": 35,  "ops": 8.0},
    {"name": "Zone Transfer",              "row": A_UC_START + 3, "tier": "Tier 1", "qtrs_after": 3, "avg_min": 90,  "ops": 5.0},
    {"name": "Mixed-Case Depalletizing",   "row": A_UC_START + 4, "tier": "Tier 1", "qtrs_after": 5, "avg_min": 80,  "ops": 6.0},
]

# Expansion pipeline: new systems signed per quarter Q1'26-Q4'30 (20 quarters)
# Q1'27-Q4'28 = 0 (no expansion activity yet), Q1'29+ = original pipeline values
PIPELINE_SYSTEMS = [0, 0, 0, 0, 0, 0, 0, 0, 1, 2, 2, 3, 4, 3, 4, 4, 5, 4, 6, 5]

# --- Pipeline Assumptions Sheet Row Map ---
PA_HDR = 1
PA_QHDR = 3
PA_QIDX = 4
PA_NEWPARTNERS = 6
PA_AVGZONES = 7
PA_NEWZONES = 8
PA_CUMZONES = 9
PA_GOLIVE = 10
PA_CUMGOLIVE = 11

# Scenario volumes [bear, base, bull, evidence_range]
SCENARIOS = [
    ["Autonomous Picking",         12, 18, 28, "10-35"],
    ["Pallet Sorting",             18, 24, 45, "15-50"],
    ["Inventory Scanning",          4,  8, 14, "2-15"],
    ["Zone Transfer",               3,  5,  8, "2-8"],
    ["Mixed-Case Depalletizing",    3,  6, 10, "2-12"],
]


def qcol(qi):
    """Quarter index (1-20) to column letter. qi=1 -> C, qi=2 -> D, ..., qi=20 -> V."""
    return get_column_letter(qi + 2)


def set_cell(ws, row, col, value, font=None, fmt=None, fill=None, alignment=None):
    """Set cell value with optional styling."""
    if isinstance(col, str):
        cell = ws[f"{col}{row}"]
    else:
        cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell


def write_quarter_headers(ws, row, start_col=3):
    """Write Q1'27..Q4'31 headers in dark blue with white bold font."""
    for i, q in enumerate(QUARTERS):
        cell = ws.cell(row=row, column=start_col + i)
        cell.value = q
        cell.font = WHITE_BOLD
        cell.fill = DARK_BLUE_FILL
        cell.alignment = Alignment(horizontal="center")


def write_quarter_indices(ws, row, start_col=3):
    """Write 1..20 quarter indices in gray small font."""
    for i in range(NUM_Q):
        cell = ws.cell(row=row, column=start_col + i)
        cell.value = i + 1
        cell.font = GRAY_SMALL
        cell.alignment = Alignment(horizontal="center")


# ============================================================
# SHEET 1: ASSUMPTIONS
# ============================================================
def build_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = "003366"

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    for c in range(3, 23):  # C-V
        ws.column_dimensions[get_column_letter(c)].width = 14

    # --- Quarter Index Reference (rows 4-6) ---
    set_cell(ws, 4, "B", "Quarter Index Reference", font=BOLD_FONT)
    write_quarter_headers(ws, 5, 3)
    write_quarter_indices(ws, 6, 3)

    # --- Anchor Distribution Partners (rows 8-26) ---
    set_cell(ws, 8, "B", "Anchor Distribution Partners", font=BOLD_FONT)

    # Row 9: System headers
    for s in SYSTEMS:
        set_cell(ws, 9, s["col"], s["name"], font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # Row 10: Number of Zones
    set_cell(ws, A_ZONES, "B", "Number of Zones")
    for s in SYSTEMS:
        set_cell(ws, A_ZONES, s["col"], s["zones"], font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # Row 11: Planning Start Quarter
    set_cell(ws, A_PLAN_Q, "B", "Planning Start Quarter")
    for s in SYSTEMS:
        set_cell(ws, A_PLAN_Q, s["col"], s["plan"], font=BLUE_FONT, fill=YELLOW_FILL)

    # Row 12: Go-Live Quarter
    set_cell(ws, A_GOLIVE_Q, "B", "Go-Live Quarter")
    for s in SYSTEMS:
        set_cell(ws, A_GOLIVE_Q, s["col"], s["golive"], font=BLUE_FONT, fill=YELLOW_FILL)

    # Row 13: License Activation Quarter
    set_cell(ws, A_LICENSE_Q, "B", "License Activation Quarter")
    for s in SYSTEMS:
        set_cell(ws, A_LICENSE_Q, s["col"], s["license"], font=BLUE_FONT, fill=YELLOW_FILL)

    # Row 14: Implementation MSRP (formula)
    set_cell(ws, A_IMPL_MSRP, "B", "Implementation MSRP")
    for s in SYSTEMS:
        set_cell(ws, A_IMPL_MSRP, s["col"],
                 f'={s["col"]}{A_ZONES}*$C${A_PERZONE_MSRP}', font=BLACK_FONT, fmt=CUR_FMT)

    # Row 15: Annual Licensing Fee (formula)
    set_cell(ws, A_LIC_FEE, "B", "Annual Licensing Fee")
    for s in SYSTEMS:
        set_cell(ws, A_LIC_FEE, s["col"],
                 f'={s["col"]}{A_ZONES}*$C${A_PERZONE_LIC}', font=BLACK_FONT, fmt=CUR_FMT)

    # Row 17: Per-Zone Implementation MSRP
    set_cell(ws, A_PERZONE_MSRP, "B", "Per-Zone Implementation MSRP")
    set_cell(ws, A_PERZONE_MSRP, "C", 350000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)

    # Row 18: Per-Zone Annual License
    set_cell(ws, A_PERZONE_LIC, "B", "Per-Zone Annual License")
    set_cell(ws, A_PERZONE_LIC, "C", 85000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)

    # Row 19: Total Anchor Zones
    set_cell(ws, 19, "B", "Total Anchor Zones")
    set_cell(ws, 19, "C", f"=C{A_ZONES}+D{A_ZONES}+E{A_ZONES}+F{A_ZONES}+G{A_ZONES}",
             font=BLACK_FONT, fmt=INT_FMT)

    # Row 20: Total Implementation Revenue
    set_cell(ws, 20, "B", "Total Implementation Revenue")
    set_cell(ws, 20, "C", f"=C{A_IMPL_MSRP}+D{A_IMPL_MSRP}+E{A_IMPL_MSRP}+F{A_IMPL_MSRP}+G{A_IMPL_MSRP}",
             font=BLACK_FONT, fmt=CUR_FMT)

    # Row 21: Total Annual Licensing
    set_cell(ws, 21, "B", "Total Annual Licensing")
    set_cell(ws, 21, "C", f"=C{A_LIC_FEE}+D{A_LIC_FEE}+E{A_LIC_FEE}+F{A_LIC_FEE}+G{A_LIC_FEE}",
             font=BLACK_FONT, fmt=CUR_FMT)

    # Row 23: Implementation Payment Terms
    set_cell(ws, 23, "B", "Implementation Payment Terms")
    set_cell(ws, 23, "C", "40/35/25", font=BLUE_FONT, fill=YELLOW_FILL)

    # Row 24-26: Payment percentages
    set_cell(ws, A_SIGNING_PCT, "B", "% at Signing")
    set_cell(ws, A_SIGNING_PCT, "C", 0.40, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_GOLIVE_PCT, "B", "% at Go-Live")
    set_cell(ws, A_GOLIVE_PCT, "C", 0.35, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_MONTH12_PCT, "B", "% at Month 12")
    set_cell(ws, A_MONTH12_PCT, "C", 0.25, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)

    # --- Zone Go-Live Ramp (rows 28-37) ---
    set_cell(ws, A_ZONE_RAMP_HDR, "B", "Zone Go-Live Ramp", font=BOLD_FONT)

    # Quarter sub-headers (Q1'27 through Q4'28 = 8 quarters)
    ramp_quarters = QUARTERS[:8]
    for i, q in enumerate(ramp_quarters):
        col = get_column_letter(3 + i)  # C through J
        set_cell(ws, A_ZONE_RAMP_QHDR, col, q, font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # Zone ramp values per partner
    for si, s in enumerate(SYSTEMS):
        r = A_ZONE_RAMP_START + si
        set_cell(ws, r, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi_idx, val in enumerate(ZONE_RAMP_DATA[si]):
            col = get_column_letter(3 + qi_idx)
            set_cell(ws, r, col, val, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # Total row
    set_cell(ws, A_ZONE_RAMP_TOTAL, "B", "  Total Zones Going Live", font=BOLD_FONT)
    for qi_idx in range(8):
        col = get_column_letter(3 + qi_idx)
        refs = "+".join(f"{col}{A_ZONE_RAMP_START + si}" for si in range(5))
        set_cell(ws, A_ZONE_RAMP_TOTAL, col, f"={refs}", font=BOLD_FONT, fmt=INT_FMT)

    # Cumulative row
    set_cell(ws, A_ZONE_RAMP_CUM, "B", "  Cumulative Zones Live", font=BOLD_FONT)
    set_cell(ws, A_ZONE_RAMP_CUM, "C", f"=C{A_ZONE_RAMP_TOTAL}", font=BOLD_FONT, fmt=INT_FMT)
    for qi_idx in range(1, 8):
        col = get_column_letter(3 + qi_idx)
        prev = get_column_letter(2 + qi_idx)
        set_cell(ws, A_ZONE_RAMP_CUM, col,
                 f"={prev}{A_ZONE_RAMP_CUM}+{col}{A_ZONE_RAMP_TOTAL}",
                 font=BOLD_FONT, fmt=INT_FMT)

    # --- Quarter Index Lookups (rows 38-41) ---
    set_cell(ws, A_QI_LOOKUP_HDR, "B", "Quarter Index Lookups", font=GRAY_BOLD)
    set_cell(ws, A_PLAN_IDX, "B", "Planning Start Index", font=GRAY_SMALL)
    set_cell(ws, A_GOLIVE_IDX, "B", "Go-Live Index", font=GRAY_SMALL)
    set_cell(ws, A_LICENSE_IDX, "B", "License Activation Index", font=GRAY_SMALL)

    for s in SYSTEMS:
        c = s["col"]
        set_cell(ws, A_PLAN_IDX, c,
                 f'=MATCH({c}{A_PLAN_Q},$C$5:$V$5,0)', font=GRAY_SMALL, fmt=INT_FMT)
        set_cell(ws, A_GOLIVE_IDX, c,
                 f'=MATCH({c}{A_GOLIVE_Q},$C$5:$V$5,0)', font=GRAY_SMALL, fmt=INT_FMT)
        set_cell(ws, A_LICENSE_IDX, c,
                 f'=MATCH({c}{A_LICENSE_Q},$C$5:$V$5,0)', font=GRAY_SMALL, fmt=INT_FMT)

    # --- Expansion Customer Pricing ---
    set_cell(ws, A_EXP_HDR, "B", "Expansion Customer Pricing", font=BOLD_FONT)
    set_cell(ws, A_EXP_IMPL_FEE, "B", "Implementation Fee (per zone)")
    set_cell(ws, A_EXP_IMPL_FEE, "C", 30000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_EXP_CONSULT, "B", "Program Consulting Fee (per partner)")
    set_cell(ws, A_EXP_CONSULT, "C", 75000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_EXP_SUB, "B", "Annual Subscription (per zone/year)")
    set_cell(ws, A_EXP_SUB, "C", 95000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_EXP_IMPL_MO, "B", "Planning + Implementation (months)")
    set_cell(ws, A_EXP_IMPL_MO, "C", 8, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_EXP_IMPL_Q, "B", "Implementation to Go-Live (quarters)")
    set_cell(ws, A_EXP_IMPL_Q, "C", f'=ROUNDUP(C{A_EXP_IMPL_MO}/3,0)', font=BLACK_FONT, fmt=INT_FMT)

    # --- Expansion Pipeline (moved to separate tab) ---
    set_cell(ws, A_PIPE_HDR, "B", "See 'Pipeline Assumptions' tab",
             font=Font(name=FONT_NAME, color="808080", italic=True))

    # --- Utilization Pricing ---
    set_cell(ws, A_UTIL_HDR, "B", "Utilization Pricing", font=BOLD_FONT)
    set_cell(ws, A_TIER1, "B", "Tier 1 $/minute")
    set_cell(ws, A_TIER1, "C", 3.50, font=BLUE_FONT, fmt='$#,##0.00', fill=YELLOW_FILL)
    set_cell(ws, A_TIER2, "B", "Tier 2 $/minute")
    set_cell(ws, A_TIER2, "C", 0.75, font=BLUE_FONT, fmt='$#,##0.00', fill=YELLOW_FILL)
    set_cell(ws, A_MAINT_PRICE, "B", "Predictive Maintenance $/robot/month")
    set_cell(ws, A_MAINT_PRICE, "C", 40.00, font=BLUE_FONT, fmt='$#,##0.00', fill=YELLOW_FILL)
    set_cell(ws, A_INSP_PRICE, "B", "Vision Inspection $/scan")
    set_cell(ws, A_INSP_PRICE, "C", 12.00, font=BLUE_FONT, fmt='$#,##0.00', fill=YELLOW_FILL)

    # --- Use Case Definitions ---
    set_cell(ws, A_UC_HDR, "B", "Use Case Definitions", font=BOLD_FONT)
    headers_uc = ["Use Case", "Tier", "Qtrs After Go-Live", "Avg Min", "Ops/Zone/Wk (Maturity)"]
    for i, h in enumerate(headers_uc):
        set_cell(ws, A_UC_COLS, get_column_letter(2 + i), h, font=BOLD_FONT)

    uc_data = [
        ("Autonomous Picking",         "Tier 1", 0, 25,  18.0),
        ("Pallet Sorting",             "Tier 2", 1, 15,  24.0),
        ("Inventory Scanning",         "Tier 1", 2, 35,   8.0),
        ("Zone Transfer",              "Tier 1", 3, 90,   5.0),
        ("Mixed-Case Depalletizing",   "Tier 1", 5, 80,   6.0),
        ("Predictive Maintenance",     "Fixed",  0, None, None),
        ("Vision Quality Inspection",  "Fixed",  1, None, None),
    ]
    for i, (name, tier, qtrs, avg_min, procs) in enumerate(uc_data):
        r = A_UC_START + i
        set_cell(ws, r, "B", name, font=BLACK_FONT)
        set_cell(ws, r, "C", tier, font=BLUE_FONT, fill=YELLOW_FILL)
        set_cell(ws, r, "D", qtrs, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
        if avg_min is not None:
            set_cell(ws, r, "E", avg_min, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
        else:
            set_cell(ws, r, "E", "\u2014", font=GRAY_SMALL)
        if procs is not None:
            set_cell(ws, r, "F", procs, font=BLUE_FONT, fmt='0.0', fill=YELLOW_FILL)
        else:
            set_cell(ws, r, "F", "\u2014", font=GRAY_SMALL)

    # [New Use Case] placeholder
    set_cell(ws, A_UC_START + 7, "B", "[New Use Case]",
             font=Font(name=FONT_NAME, color="808080", italic=True))
    set_cell(ws, A_UC_START + 7, "C", "Tier 1", font=BLUE_FONT, fill=YELLOW_FILL)
    set_cell(ws, A_UC_START + 7, "D", 0, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_UC_START + 7, "E", 0, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_UC_START + 7, "F", 0, font=BLUE_FONT, fmt='0.0', fill=YELLOW_FILL)

    # --- Hockey Stick Ramp ---
    set_cell(ws, A_RAMP_HDR, "B", "Hockey Stick Ramp", font=BOLD_FONT)
    ramp_values = [0.12, 0.18, 0.28, 0.42, 0.58, 0.75, 0.88, 1.00]
    for i, val in enumerate(ramp_values):
        r = A_RAMP_START + i
        set_cell(ws, r, "B", f"Q{i+1}")
        set_cell(ws, r, "C", val, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)

    # --- Site Volume Decay ---
    set_cell(ws, A_DECAY_HDR, "B", "Site Volume Decay", font=BOLD_FONT)
    decay_values = [1.00, 0.95, 0.88, 0.80, 0.72, 0.65]
    for i, val in enumerate(decay_values):
        r = A_DECAY_ZONE_START + i
        set_cell(ws, r, "B", f"Zone {i+1}")
        set_cell(ws, r, "C", val, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)

    set_cell(ws, A_DECAY_6, "B", "Blended Average (6 zones)")
    set_cell(ws, A_DECAY_6, "C",
             f"=AVERAGE(C{A_DECAY_ZONE_START}:C{A_DECAY_ZONE_START+5})", font=BLACK_FONT, fmt=PCT_FMT)
    set_cell(ws, A_DECAY_5, "B", "Blended Average (5 zones)")
    set_cell(ws, A_DECAY_5, "C",
             f"=AVERAGE(C{A_DECAY_ZONE_START}:C{A_DECAY_ZONE_START+4})", font=BLACK_FONT, fmt=PCT_FMT)
    set_cell(ws, A_DECAY_4, "B", "Blended Average (4 zones)")
    set_cell(ws, A_DECAY_4, "C",
             f"=AVERAGE(C{A_DECAY_ZONE_START}:C{A_DECAY_ZONE_START+3})", font=BLACK_FONT, fmt=PCT_FMT)
    set_cell(ws, A_DECAY_3, "B", "Blended Average (3 zones)")
    set_cell(ws, A_DECAY_3, "C",
             f"=AVERAGE(C{A_DECAY_ZONE_START}:C{A_DECAY_ZONE_START+2})", font=BLACK_FONT, fmt=PCT_FMT)
    set_cell(ws, A_DECAY_2, "B", "Blended Average (2 zones)")
    set_cell(ws, A_DECAY_2, "C",
             f"=AVERAGE(C{A_DECAY_ZONE_START}:C{A_DECAY_ZONE_START+1})", font=BLACK_FONT, fmt=PCT_FMT)

    # --- Predictive Maintenance ---
    set_cell(ws, A_MAINT_HDR, "B", "Predictive Maintenance", font=BOLD_FONT)
    set_cell(ws, A_ROBOTS, "B", "Robots per Zone")
    set_cell(ws, A_ROBOTS, "C", 80, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_MAINT_RAMP_START, "B", "Maintenance Ramp Start")
    set_cell(ws, A_MAINT_RAMP_START, "C", 0.40, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_MAINT_RAMP_Q, "B", "Quarters to Full Maintenance")
    set_cell(ws, A_MAINT_RAMP_Q, "C", 5, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # --- Vision Quality Inspection ---
    set_cell(ws, A_INSP_HDR, "B", "Vision Quality Inspection", font=BOLD_FONT)
    set_cell(ws, A_SCANS, "B", "Scans per Zone per Month")
    set_cell(ws, A_SCANS, "C", 250, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # --- Scenario Volumes ---
    set_cell(ws, A_SCENARIO_HDR, "B", "Scenario Volumes (ops/zone/wk at maturity)", font=BOLD_FONT)
    scenario_headers = ["Use Case", "Bear", "Base", "Bull", "Evidence Range"]
    for i, h in enumerate(scenario_headers):
        set_cell(ws, A_SCENARIO_COLS, get_column_letter(2 + i), h, font=BOLD_FONT)

    for i, (name, bear, base, bull, ev_range) in enumerate(SCENARIOS):
        r = A_SCENARIO_START + i
        set_cell(ws, r, "B", name, font=BLACK_FONT)
        set_cell(ws, r, "C", bear, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
        set_cell(ws, r, "D", base, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
        set_cell(ws, r, "E", bull, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
        set_cell(ws, r, "F", ev_range, font=BLUE_FONT, fill=YELLOW_FILL)

    # --- Site Cost Assumptions ---
    set_cell(ws, A_COST_HDR, "B", "Site Cost Assumptions", font=BOLD_FONT)
    set_cell(ws, A_NRC_FIRST, "B", "Non-Recurring Cost (first zone)")
    set_cell(ws, A_NRC_FIRST, "C", 62000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_MRC, "B", "Monthly Recurring Cost (per zone)")
    set_cell(ws, A_MRC, "C", 18000, font=BLUE_FONT, fmt=CUR_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_DISC, "B", "Additional Zone Discount")
    set_cell(ws, A_DISC, "C", 0.35, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_ARC, "B", "Annual Recurring Cost (per site)")
    set_cell(ws, A_ARC, "C", f"=C{A_MRC}*12", font=BLACK_FONT, fmt=CUR_FMT)
    set_cell(ws, A_QRC, "B", "Quarterly Recurring Cost (per site)")
    set_cell(ws, A_QRC, "C", f"=C{A_MRC}*3", font=BLACK_FONT, fmt=CUR_FMT)
    set_cell(ws, A_NRC_ADDL, "B", "NRC per Additional Zone")
    set_cell(ws, A_NRC_ADDL, "C", f"=C{A_NRC_FIRST}*(1-C{A_DISC})", font=BLACK_FONT, fmt=CUR_FMT)
    set_cell(ws, A_QRC_ADDL, "B", "Quarterly Recurring per Add'l Zone")
    set_cell(ws, A_QRC_ADDL, "C", f"=C{A_QRC}*(1-C{A_DISC})", font=BLACK_FONT, fmt=CUR_FMT)
    set_cell(ws, A_BLEND_NRC, "B", "Blended NRC per Zone (3-zone avg)")
    set_cell(ws, A_BLEND_NRC, "C",
             f"=(C{A_NRC_FIRST}+('Pipeline Assumptions'!C{PA_AVGZONES}-1)*C{A_NRC_ADDL})/'Pipeline Assumptions'!C{PA_AVGZONES}",
             font=BLACK_FONT, fmt=CUR_FMT)
    set_cell(ws, A_BLEND_QRC, "B", "Blended Qtly Recurring per Zone (3-zone avg)")
    set_cell(ws, A_BLEND_QRC, "C",
             f"=(C{A_QRC}+('Pipeline Assumptions'!C{PA_AVGZONES}-1)*C{A_QRC_ADDL})/'Pipeline Assumptions'!C{PA_AVGZONES}",
             font=BLACK_FONT, fmt=CUR_FMT)

    # --- Operating Expense Assumptions ---
    set_cell(ws, A_OPEX_HDR, "B", "Operating Expense Assumptions", font=BOLD_FONT)
    set_cell(ws, A_GA_PCT, "B", "G&A OpEx (% of Revenue)")
    set_cell(ws, A_GA_PCT, "C", 0.12, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_RD_PCT, "B", "R&D OpEx (% of Revenue)")
    set_cell(ws, A_RD_PCT, "C", 0.25, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_AMORT_YRS, "B", "CapEx Amortization Period (years)")
    set_cell(ws, A_AMORT_YRS, "C", 5, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)
    set_cell(ws, A_EBITDA_MULT, "B", "EBITDA Terminal Multiple")
    set_cell(ws, A_EBITDA_MULT, "C", 12, font=BLUE_FONT, fmt=MULT_FMT, fill=YELLOW_FILL)

    # --- Sources & Notes ---
    set_cell(ws, A_SOURCES_NOTE, "B",
             "See Sources & Notes section below for documentation of all assumptions",
             font=Font(name=FONT_NAME, color="808080", italic=True))

    with open("v3_sources_notes.json", "r", encoding="utf-8") as f:
        sources = json.load(f)

    target_row = A_SOURCES_START
    for entry in sources:
        b_val = entry.get("B", "")
        if b_val:
            b_val = re.sub(r'\[\d+\]\[?[a-z]?\]?', '', b_val).strip()

        c_val = entry.get("C", "")
        d_val = entry.get("D", "")
        e_val = entry.get("E", "")
        f_val = entry.get("F", "")

        is_header = b_val and not c_val and not d_val and not e_val and not f_val
        if is_header:
            set_cell(ws, target_row, "B", b_val, font=BOLD_FONT)
        else:
            set_cell(ws, target_row, "B", b_val, font=BLACK_FONT)
            if c_val:
                set_cell(ws, target_row, "C", c_val, font=BLACK_FONT)
            if d_val:
                set_cell(ws, target_row, "D", d_val, font=BLACK_FONT)
            if e_val:
                set_cell(ws, target_row, "E", e_val, font=BLACK_FONT)
            if f_val:
                set_cell(ws, target_row, "F", f_val, font=BLACK_FONT)
        target_row += 1

    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60

    return ws


# ============================================================
# SHEET 1b: PIPELINE ASSUMPTIONS
# ============================================================
def build_pipeline_assumptions(wb):
    ws = wb.create_sheet("Pipeline Assumptions")
    ws.sheet_properties.tabColor = "006633"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    for c in range(3, 23):  # C-V
        ws.column_dimensions[get_column_letter(c)].width = 10

    # Header
    set_cell(ws, PA_HDR, "B", "Expansion New Partner Pipeline",
             font=Font(name=FONT_NAME, bold=True, size=14))

    # Quarter headers + indices
    write_quarter_headers(ws, PA_QHDR, 3)
    write_quarter_indices(ws, PA_QIDX, 3)

    # Row PA_NEWPARTNERS: New Partners Signed (editable input)
    set_cell(ws, PA_NEWPARTNERS, "B", "New Partners Signed")
    for i, val in enumerate(PIPELINE_SYSTEMS):
        col = get_column_letter(3 + i)
        set_cell(ws, PA_NEWPARTNERS, col, val, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # Row PA_AVGZONES: Avg Zones per Partner (editable input, single value)
    set_cell(ws, PA_AVGZONES, "B", "Avg Zones per Partner")
    set_cell(ws, PA_AVGZONES, "C", 3, font=BLUE_FONT, fmt=INT_FMT, fill=YELLOW_FILL)

    # Row PA_NEWZONES: New Zones This Quarter (formula)
    set_cell(ws, PA_NEWZONES, "B", "New Zones This Quarter")
    for i in range(NUM_Q):
        col = get_column_letter(3 + i)
        set_cell(ws, PA_NEWZONES, col,
                 f'={col}{PA_NEWPARTNERS}*$C${PA_AVGZONES}', font=BLACK_FONT, fmt=INT_FMT)

    # Row PA_CUMZONES: Cumulative New Zones (running sum)
    set_cell(ws, PA_CUMZONES, "B", "Cumulative New Zones")
    set_cell(ws, PA_CUMZONES, "C", f'=C{PA_NEWZONES}', font=BLACK_FONT, fmt=INT_FMT)
    for i in range(1, NUM_Q):
        col = get_column_letter(3 + i)
        prev = get_column_letter(2 + i)
        set_cell(ws, PA_CUMZONES, col,
                 f'={prev}{PA_CUMZONES}+{col}{PA_NEWZONES}', font=BLACK_FONT, fmt=INT_FMT)

    # Row PA_GOLIVE: Go-Live Zones This Quarter (offset by impl quarters from Assumptions)
    set_cell(ws, PA_GOLIVE, "B", "Go-Live Zones This Quarter")
    for i in range(NUM_Q):
        col = get_column_letter(3 + i)
        set_cell(ws, PA_GOLIVE, col,
                 f'=IF(COLUMN({col}{PA_NEWZONES})-COLUMN($C${PA_NEWZONES})>=Assumptions!$C${A_EXP_IMPL_Q},'
                 f'INDEX($C${PA_NEWZONES}:$V${PA_NEWZONES},1,'
                 f'COLUMN({col}{PA_NEWZONES})-COLUMN($C${PA_NEWZONES})+1-Assumptions!$C${A_EXP_IMPL_Q}),0)',
                 font=BLACK_FONT, fmt=INT_FMT)

    # Row PA_CUMGOLIVE: Cumulative Go-Live Zones (running sum)
    set_cell(ws, PA_CUMGOLIVE, "B", "Cumulative Go-Live Zones")
    set_cell(ws, PA_CUMGOLIVE, "C", f'=C{PA_GOLIVE}', font=BLACK_FONT, fmt=INT_FMT)
    for i in range(1, NUM_Q):
        col = get_column_letter(3 + i)
        prev = get_column_letter(2 + i)
        set_cell(ws, PA_CUMGOLIVE, col,
                 f'={prev}{PA_CUMGOLIVE}+{col}{PA_GOLIVE}', font=BLACK_FONT, fmt=INT_FMT)

    return ws


# ============================================================
# SHEET 2: ANCHOR REVENUE
# ============================================================
def build_anchor(wb):
    ws = wb.create_sheet("Anchor Revenue")
    ws.sheet_properties.tabColor = "003366"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    for c in range(3, 23):
        ws.column_dimensions[get_column_letter(c)].width = 14

    set_cell(ws, 1, "B", "Anchor Revenue", font=Font(name=FONT_NAME, bold=True, size=14))
    write_quarter_headers(ws, 3, 3)
    write_quarter_indices(ws, 4, 3)

    # ---- CUMULATIVE ZONES LIVE HELPER (per system, all 20 quarters) ----
    row = 6
    set_cell(ws, row, "B", "Cumulative Zones Live (helper)", font=GRAY_BOLD)
    row += 1

    cum_zone_rows = []  # track row number per system
    for si, s in enumerate(SYSTEMS):
        ramp_row = A_ZONE_RAMP_START + si
        set_cell(ws, row, "B", f"  {s['name']}", font=GRAY_SMALL)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            # For qi 1-8: SUM from C to the column corresponding to qi
            # For qi > 8: SUM entire ramp row (all zones deployed)
            end_col_idx = min(qi + 2, 10)  # J = column 10, max ramp column
            end_col = get_column_letter(end_col_idx)
            formula = f"=SUM(Assumptions!$C${ramp_row}:{end_col}${ramp_row})"
            set_cell(ws, row, col, formula, font=GRAY_SMALL, fmt=INT_FMT)
        cum_zone_rows.append(row)
        row += 1
    row += 1  # blank

    # ---- SECTION 1: IMPLEMENTATION REVENUE ----
    set_cell(ws, row, "B", "Implementation Revenue", font=BOLD_FONT)
    row += 1

    impl_rows = []
    for si, s in enumerate(SYSTEMS):
        sys_col = s["col"]
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            formula = (
                f'=IF({qi}=Assumptions!${sys_col}${A_PLAN_IDX},'
                f'Assumptions!${sys_col}${A_IMPL_MSRP}*Assumptions!$C${A_SIGNING_PCT},0)'
                f'+IF({qi}=Assumptions!${sys_col}${A_GOLIVE_IDX},'
                f'Assumptions!${sys_col}${A_IMPL_MSRP}*Assumptions!$C${A_GOLIVE_PCT},0)'
                f'+IF({qi}=Assumptions!${sys_col}${A_GOLIVE_IDX}+4,'
                f'Assumptions!${sys_col}${A_IMPL_MSRP}*Assumptions!$C${A_MONTH12_PCT},0)'
            )
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        impl_rows.append(row)
        row += 1
        if si < len(SYSTEMS) - 1:
            row += 1  # blank row between systems

    row += 1
    impl_total_row = row
    set_cell(ws, row, "B", "Total Implementation", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in impl_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 2: NON-GAAP IMPLEMENTATION + CONSULTING FEE ----
    row += 2
    set_cell(ws, row, "B", "NON-GAAP IMPL + CONSULTING FEE", font=BOLD_FONT)
    row += 1

    nongaap_impl_rows = []
    for si, s in enumerate(SYSTEMS):
        sys_col = s["col"]
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            # One-time fee at planning start: $30k * zones + $75k consulting
            formula = (
                f'=IF({qi}=Assumptions!${sys_col}${A_PLAN_IDX},'
                f'Assumptions!${sys_col}${A_ZONES}*Assumptions!$C${A_EXP_IMPL_FEE}'
                f'+Assumptions!$C${A_EXP_CONSULT},0)'
            )
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        nongaap_impl_rows.append(row)
        row += 1

    row += 1
    nongaap_impl_total_row = row
    set_cell(ws, row, "B", "Total Non-GAAP Impl + Consulting", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in nongaap_impl_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 3: NON-GAAP SUBSCRIPTION REVENUE (pre-activation licensing) ----
    row += 2
    set_cell(ws, row, "B", "NON-GAAP SUBSCRIPTION REVENUE", font=BOLD_FONT)
    row += 1

    nongaap_rows = []
    for si, s in enumerate(SYSTEMS):
        sys_col = s["col"]
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            cum_ref = f"{col}{cum_zone_rows[si]}"
            # Revenue when zones are live but before license activation
            formula = (
                f'=IF(AND({cum_ref}>0,{qi}<Assumptions!${sys_col}${A_LICENSE_IDX}),'
                f'{cum_ref}*Assumptions!$C${A_PERZONE_LIC}/4,0)'
            )
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        nongaap_rows.append(row)
        row += 1

    row += 1
    nongaap_total_row = row
    set_cell(ws, row, "B", "Total Non-GAAP Subscription", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in nongaap_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 4: ZONE LICENSING FEE (GAAP — starts at license activation) ----
    row += 2
    set_cell(ws, row, "B", "Zone Licensing Fee", font=BOLD_FONT)
    row += 1

    lic_rows = []
    for si, s in enumerate(SYSTEMS):
        sys_col = s["col"]
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            formula = (
                f'=IF({qi}>=Assumptions!${sys_col}${A_LICENSE_IDX},'
                f'Assumptions!${sys_col}${A_LIC_FEE}/4,0)'
            )
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        lic_rows.append(row)
        row += 1

    row += 1
    lic_total_row = row
    set_cell(ws, row, "B", "Total Licensing", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in lic_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 5: OPERATIONS REVENUE (cohort-based) ----
    row += 2
    set_cell(ws, row, "B", "Operations Revenue", font=BOLD_FONT)
    row += 1

    uc_subtotal_rows = []
    for uc in USE_CASES:
        set_cell(ws, row, "B", f"  {uc['name']}", font=Font(name=FONT_NAME, bold=True, italic=True))
        row += 1

        uc_sys_rows = []
        for si, s in enumerate(SYSTEMS):
            sys_col = s["col"]
            set_cell(ws, row, "B", f"    {s['name']}", font=BLACK_FONT)

            price_ref = f"Assumptions!$C${A_TIER1}" if uc["tier"] == "Tier 1" else f"Assumptions!$C${A_TIER2}"
            uc_row = uc["row"]
            qtrs_after = uc["qtrs_after"]
            # Decay based on total zone count for this system
            decay_formula = (
                f'IF(Assumptions!${sys_col}${A_ZONES}>=6,Assumptions!$C${A_DECAY_6},'
                f'IF(Assumptions!${sys_col}${A_ZONES}>=5,Assumptions!$C${A_DECAY_5},'
                f'IF(Assumptions!${sys_col}${A_ZONES}>=4,Assumptions!$C${A_DECAY_4},'
                f'IF(Assumptions!${sys_col}${A_ZONES}>=3,Assumptions!$C${A_DECAY_3},'
                f'IF(Assumptions!${sys_col}${A_ZONES}>=2,Assumptions!$C${A_DECAY_2},1)))))'
            )

            for qi in range(1, NUM_Q + 1):
                col = qcol(qi)
                # Cohort-based: sum across 8 ramp grid quarters
                terms = []
                for j in range(1, 9):  # j=1..8 = Q1'26..Q4'27
                    elapsed = qi - j - qtrs_after
                    if elapsed < 1:
                        continue  # cohort not yet activated
                    ramp_col_j = get_column_letter(j + 2)  # C=3, so j=1->C, j=8->J
                    ramp_zones_ref = f"Assumptions!{ramp_col_j}${A_ZONE_RAMP_START + si}"
                    if elapsed >= 8:
                        ramp_part = "1"
                    else:
                        ramp_part = f"INDEX(Assumptions!{A_RAMP_RANGE},{elapsed})"
                    term = (
                        f"{ramp_zones_ref}*{ramp_part}"
                        f"*Assumptions!$F${uc_row}*13"
                        f"*Assumptions!$E${uc_row}"
                        f"*{decay_formula}"
                        f"*{price_ref}"
                    )
                    terms.append(term)

                if terms:
                    set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
                else:
                    set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)

            uc_sys_rows.append(row)
            row += 1

        set_cell(ws, row, "B", f"  Subtotal {uc['name']}", font=BOLD_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            refs = "+".join(f"{col}{r}" for r in uc_sys_rows)
            set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)
        uc_subtotal_rows.append(row)
        row += 1
        row += 1  # blank row between use cases

    proc_total_row = row
    set_cell(ws, row, "B", "Total Operations Revenue", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in uc_subtotal_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 6: PREDICTIVE MAINTENANCE (cohort-based) ----
    row += 2
    set_cell(ws, row, "B", "Predictive Maintenance Revenue", font=BOLD_FONT)
    row += 1

    mon_rows = []
    for si, s in enumerate(SYSTEMS):
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            terms = []
            for j in range(1, 9):
                elapsed = qi - j
                if elapsed < 0:
                    continue
                ramp_col_j = get_column_letter(j + 2)
                ramp_zones_ref = f"Assumptions!{ramp_col_j}${A_ZONE_RAMP_START + si}"
                ramp_part = (
                    f"MIN(1,Assumptions!$C${A_MAINT_RAMP_START}"
                    f"+({elapsed}/Assumptions!$C${A_MAINT_RAMP_Q})"
                    f"*(1-Assumptions!$C${A_MAINT_RAMP_START}))"
                )
                term = (
                    f"{ramp_zones_ref}*{ramp_part}"
                    f"*Assumptions!$C${A_ROBOTS}*Assumptions!$C${A_MAINT_PRICE}*3"
                )
                terms.append(term)

            if terms:
                set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
            else:
                set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
        mon_rows.append(row)
        row += 1

    row += 1
    mon_total_row = row
    set_cell(ws, row, "B", "Total Maintenance Revenue", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in mon_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 7: VISION QUALITY INSPECTION (cohort-based) ----
    row += 2
    set_cell(ws, row, "B", "VISION QUALITY INSPECTION REVENUE", font=BOLD_FONT)
    row += 1

    img_rows = []
    for si, s in enumerate(SYSTEMS):
        set_cell(ws, row, "B", f"  {s['name']}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            terms = []
            for j in range(1, 9):
                elapsed = qi - j - 1  # inspection starts go-live + 1
                if elapsed < 1:
                    continue
                ramp_col_j = get_column_letter(j + 2)
                ramp_zones_ref = f"Assumptions!{ramp_col_j}${A_ZONE_RAMP_START + si}"
                if elapsed >= 8:
                    ramp_part = "1"
                else:
                    ramp_part = f"INDEX(Assumptions!{A_RAMP_RANGE},{elapsed})"
                term = (
                    f"{ramp_zones_ref}*{ramp_part}"
                    f"*Assumptions!$C${A_SCANS}*3*Assumptions!$C${A_INSP_PRICE}"
                )
                terms.append(term)

            if terms:
                set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
            else:
                set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
        img_rows.append(row)
        row += 1

    row += 1
    img_total_row = row
    set_cell(ws, row, "B", "Total Inspection Revenue", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in img_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SUMMARY SECTION ----
    row += 2
    set_cell(ws, row, "B", "Summary", font=BOLD_FONT)
    row += 1

    summary_labels = [
        ("Implementation Revenue", impl_total_row),
        ("Non-GAAP Impl + Consulting", nongaap_impl_total_row),
        ("Non-GAAP Subscription Revenue", nongaap_total_row),
        ("Zone Licensing Fee", lic_total_row),
        ("Operations Revenue", proc_total_row),
        ("Predictive Maintenance", mon_total_row),
        ("Vision Quality Inspection", img_total_row),
    ]
    summary_rows = []
    for label, src_row in summary_labels:
        set_cell(ws, row, "B", f"  {label}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            set_cell(ws, row, col, f"={col}{src_row}", font=BLACK_FONT, fmt=CUR_FMT)
        summary_rows.append(row)
        row += 1

    # GAAP total (excludes Non-GAAP subscription = summary_rows[1])
    row += 1
    anchor_total_row = row
    set_cell(ws, row, "B", "Total Anchor Revenue", font=BOLD_FONT)
    gaap_rows = [summary_rows[0]] + summary_rows[3:]  # skip Non-GAAP (indices 1,2)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in gaap_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # Total incl. Non-GAAP
    row += 1
    anchor_total_incl_nongaap_row = row
    set_cell(ws, row, "B", "Total Anchor incl. Non-GAAP", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col,
                 f"={col}{anchor_total_row}+{col}{summary_rows[1]}+{col}{summary_rows[2]}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # ---- ANNUAL SUMMARY ----
    row += 2
    set_cell(ws, row, "B", "Annual Summary", font=BOLD_FONT)
    row += 1

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, row, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))
    row += 1

    annual_src_labels = [
        ("Implementation", impl_total_row),
        ("Non-GAAP Impl + Consulting", nongaap_impl_total_row),
        ("Non-GAAP Subscription", nongaap_total_row),
        ("Licensing", lic_total_row),
        ("Operations", proc_total_row),
        ("Maintenance", mon_total_row),
        ("Inspection", img_total_row),
    ]
    annual_rows = []
    for label, src_row in annual_src_labels:
        set_cell(ws, row, "B", f"  {label}", font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            q_start = 1 + 4 * yi
            q_cols = [qcol(q_start + q) for q in range(4)]
            formula = "=" + "+".join(f"'{ws.title}'!{c}{src_row}" for c in q_cols)
            set_cell(ws, row, yr_col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        annual_rows.append(row)
        row += 1

    row += 1
    ann_total_row = row
    set_cell(ws, row, "B", "Total Annual Revenue", font=BOLD_FONT)
    gaap_annual = [annual_rows[0]] + annual_rows[3:]  # skip Non-GAAP (indices 1,2)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        refs = "+".join(f"{yr_col}{r}" for r in gaap_annual)
        set_cell(ws, row, yr_col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 1
    ann_total_incl_nongaap_row = row
    set_cell(ws, row, "B", "Total incl. Non-GAAP", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, row, yr_col,
                 f"={yr_col}{ann_total_row}+{yr_col}{annual_rows[1]}+{yr_col}{annual_rows[2]}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    row += 1
    set_cell(ws, row, "B", "Cumulative Revenue", font=BOLD_FONT)
    set_cell(ws, row, "C", f"=C{ann_total_row}", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, row, yr_col, f"={prev_col}{row}+{yr_col}{ann_total_row}", font=BOLD_FONT, fmt=CUR_FMT)
    cum_row = row

    # Store key rows for cross-sheet references
    ws._anchor_total_row = anchor_total_row
    ws._anchor_total_incl_nongaap_row = anchor_total_incl_nongaap_row
    ws._impl_total_row = impl_total_row
    ws._nongaap_total_row = nongaap_total_row
    ws._lic_total_row = lic_total_row
    ws._proc_total_row = proc_total_row
    ws._mon_total_row = mon_total_row
    ws._img_total_row = img_total_row
    ws._ann_total_row = ann_total_row
    ws._ann_total_incl_nongaap_row = ann_total_incl_nongaap_row
    ws._nongaap_impl_total_row = nongaap_impl_total_row
    ws._annual_rows = annual_rows  # [impl, nongaap_impl, nongaap_sub, lic, proc, mon, img]
    ws._cum_row = cum_row
    ws._cum_zone_rows = cum_zone_rows  # per-partner cumulative zone helper rows

    return ws


# ============================================================
# SHEET 3: EXPANSION PIPELINE
# ============================================================
def build_expansion(wb):
    ws = wb.create_sheet("Expansion Pipeline")
    ws.sheet_properties.tabColor = "006633"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 23):
        ws.column_dimensions[get_column_letter(c)].width = 14

    set_cell(ws, 1, "B", "Expansion Pipeline Revenue", font=Font(name=FONT_NAME, bold=True, size=14))
    write_quarter_headers(ws, 3, 3)
    write_quarter_indices(ws, 4, 3)

    # ---- SECTION 1: NEW SYSTEM DEPLOYMENT ----
    row = 6
    set_cell(ws, row, "B", "New Partner Deployment", font=BOLD_FONT)
    row += 1

    # Row 7: New Systems Signed — pull from Pipeline Assumptions tab
    new_sys_row = row
    set_cell(ws, row, "B", "New Partners Signed", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"='Pipeline Assumptions'!{col}{PA_NEWPARTNERS}",
                 font=BLACK_FONT, fmt=INT_FMT)
    row += 1

    # Row 8: New Zones Signed
    new_zones_row = row
    set_cell(ws, row, "B", "New Zones Signed", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{new_sys_row}*'Pipeline Assumptions'!$C${PA_AVGZONES}",
                 font=BLACK_FONT, fmt=INT_FMT)
    row += 1

    # Row 9: Go-Live Zones (new this quarter) = zones signed impl_quarters ago
    golive_zones_row = row
    set_cell(ws, row, "B", "Go-Live Zones (new this quarter)", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        formula = (
            f'=IF(COLUMN({col}{new_zones_row})-COLUMN($C${new_zones_row})>=Assumptions!$C${A_EXP_IMPL_Q},'
            f'INDEX($C${new_zones_row}:$V${new_zones_row},1,'
            f'COLUMN({col}{new_zones_row})-COLUMN($C${new_zones_row})+1-Assumptions!$C${A_EXP_IMPL_Q}),'
            f'0)'
        )
        set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=INT_FMT)
    row += 1

    # Row 10: Cumulative Signed Zones
    cum_signed_row = row
    set_cell(ws, row, "B", "Cumulative Signed Zones", font=BLACK_FONT)
    set_cell(ws, row, "C", f"=C{new_zones_row}", font=BLACK_FONT, fmt=INT_FMT)
    for qi in range(2, NUM_Q + 1):
        col = qcol(qi)
        prev = qcol(qi - 1)
        set_cell(ws, row, col, f"={prev}{row}+{col}{new_zones_row}", font=BLACK_FONT, fmt=INT_FMT)
    row += 1

    # Row 11: Cumulative Go-Live Zones
    cum_golive_row = row
    set_cell(ws, row, "B", "Cumulative Go-Live Zones", font=BLACK_FONT)
    set_cell(ws, row, "C", f"=C{golive_zones_row}", font=BLACK_FONT, fmt=INT_FMT)
    for qi in range(2, NUM_Q + 1):
        col = qcol(qi)
        prev = qcol(qi - 1)
        set_cell(ws, row, col, f"={prev}{row}+{col}{golive_zones_row}", font=BLACK_FONT, fmt=INT_FMT)
    row += 1

    # ---- SECTION 2: CONSULTING + IMPLEMENTATION FEES ----
    row += 1
    set_cell(ws, row, "B", "Consulting + Implementation Fees", font=BOLD_FONT)
    row += 1
    consult_impl_row = row
    set_cell(ws, row, "B", "One-time Revenue at Signing", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        formula = (
            f'={col}{new_sys_row}*Assumptions!$C${A_EXP_CONSULT}'
            f'+{col}{new_zones_row}*Assumptions!$C${A_EXP_IMPL_FEE}'
        )
        set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    # ---- SECTION 3: ANNUAL SUBSCRIPTION ----
    row += 1
    set_cell(ws, row, "B", "Annual Subscription", font=BOLD_FONT)
    row += 1
    sub_row = row
    set_cell(ws, row, "B", "Subscription Revenue", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        formula = f'={col}{cum_signed_row}*Assumptions!$C${A_EXP_SUB}/4'
        set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    # ---- SECTION 4: UTILIZATION REVENUE ----
    row += 1
    set_cell(ws, row, "B", "Utilization Revenue", font=BOLD_FONT)
    row += 1

    # Helper rows: per-zone quarterly mature revenue
    set_cell(ws, row, "B", "Per-Zone Quarterly Operations Rev (maturity)", font=GRAY_SMALL)
    proc_helper_row = row
    proc_terms = []
    for uc in USE_CASES:
        price_ref = f"Assumptions!$C${A_TIER1}" if uc["tier"] == "Tier 1" else f"Assumptions!$C${A_TIER2}"
        proc_terms.append(f"Assumptions!$F${uc['row']}*13*Assumptions!$E${uc['row']}*{price_ref}")
    set_cell(ws, row, "C",
             f"=({'+'.join(proc_terms)})*Assumptions!$C${A_DECAY_3}",
             font=GRAY_SMALL, fmt=CUR_FMT)
    row += 1

    set_cell(ws, row, "B", "Per-Zone Quarterly Maintenance Rev (maturity)", font=GRAY_SMALL)
    mon_helper_row = row
    set_cell(ws, row, "C",
             f"=Assumptions!$C${A_ROBOTS}*Assumptions!$C${A_MAINT_PRICE}*3",
             font=GRAY_SMALL, fmt=CUR_FMT)
    row += 1

    set_cell(ws, row, "B", "Per-Zone Quarterly Inspection Rev (maturity)", font=GRAY_SMALL)
    img_helper_row = row
    set_cell(ws, row, "C",
             f"=Assumptions!$C${A_SCANS}*3*Assumptions!$C${A_INSP_PRICE}",
             font=GRAY_SMALL, fmt=CUR_FMT)
    row += 1
    row += 1

    # Operations Revenue with cohort-by-cohort ramp
    set_cell(ws, row, "B", "  Operations Revenue", font=BLACK_FONT)
    util_proc_row = row
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        # For each quarter, sum across all prior go-live cohorts:
        # go_live_zones_in_quarter_j * ramp(qi-j) * per_zone_mature_ops_rev
        # We look back through all possible prior quarters
        terms = []
        for j in range(1, NUM_Q + 1):
            if j > qi:
                break
            j_col = qcol(j)
            elapsed = qi - j
            if elapsed < 1:
                continue
            if elapsed >= 8:
                ramp_part = "1"
            else:
                ramp_part = f"INDEX(Assumptions!{A_RAMP_RANGE},{elapsed})"
            terms.append(f"{j_col}{golive_zones_row}*{ramp_part}*$C${proc_helper_row}")

        if terms:
            set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
        else:
            set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    # Maintenance Revenue with cohort ramp (linear ramp: 40% to 100% over 5 quarters)
    set_cell(ws, row, "B", "  Maintenance Revenue", font=BLACK_FONT)
    util_mon_row = row
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        terms = []
        for j in range(1, NUM_Q + 1):
            if j > qi:
                break
            j_col = qcol(j)
            elapsed = qi - j
            if elapsed < 0:
                continue
            ramp_part = (
                f"MIN(1,Assumptions!$C${A_MAINT_RAMP_START}"
                f"+({elapsed}/Assumptions!$C${A_MAINT_RAMP_Q})"
                f"*(1-Assumptions!$C${A_MAINT_RAMP_START}))"
            )
            terms.append(f"{j_col}{golive_zones_row}*{ramp_part}*$C${mon_helper_row}")

        if terms:
            set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
        else:
            set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    # Inspection Revenue (starts at go-live + 1 quarter, uses hockey stick ramp)
    set_cell(ws, row, "B", "  Inspection Revenue", font=BLACK_FONT)
    util_img_row = row
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        terms = []
        for j in range(1, NUM_Q + 1):
            if j > qi:
                break
            j_col = qcol(j)
            elapsed = qi - j - 1  # inspection starts 1 quarter after go-live
            if elapsed < 1:
                continue
            if elapsed >= 8:
                ramp_part = "1"
            else:
                ramp_part = f"INDEX(Assumptions!{A_RAMP_RANGE},{elapsed})"
            terms.append(f"{j_col}{golive_zones_row}*{ramp_part}*$C${img_helper_row}")

        if terms:
            set_cell(ws, row, col, "=" + "+".join(terms), font=BLACK_FONT, fmt=CUR_FMT)
        else:
            set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    row += 1
    util_total_row = row
    set_cell(ws, row, "B", "Total Utilization", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col,
                 f"={col}{util_proc_row}+{col}{util_mon_row}+{col}{util_img_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SUMMARY ----
    row += 2
    set_cell(ws, row, "B", "Summary", font=BOLD_FONT)
    row += 1

    summary_items = [
        ("Consulting + Implementation", consult_impl_row),
        ("Annual Subscription", sub_row),
        ("Utilization Revenue", util_total_row),
    ]
    exp_summary_rows = []
    for label, src_row in summary_items:
        set_cell(ws, row, "B", f"  {label}", font=BLACK_FONT)
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            set_cell(ws, row, col, f"={col}{src_row}", font=BLACK_FONT, fmt=CUR_FMT)
        exp_summary_rows.append(row)
        row += 1

    row += 1
    exp_total_row = row
    set_cell(ws, row, "B", "Total Expansion Revenue", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in exp_summary_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- ANNUAL SUMMARY ----
    row += 2
    set_cell(ws, row, "B", "Annual Summary", font=BOLD_FONT)
    row += 1

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, row, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))
    row += 1

    ann_src = [
        ("Consulting + Implementation", consult_impl_row),
        ("Annual Subscription", sub_row),
        ("Utilization Revenue", util_total_row),
    ]
    exp_annual_rows = []
    for label, src_row in ann_src:
        set_cell(ws, row, "B", f"  {label}", font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            q_start = 1 + 4 * yi
            q_cols = [qcol(q_start + q) for q in range(4)]
            formula = "=" + "+".join(f"'{ws.title}'!{c}{src_row}" for c in q_cols)
            set_cell(ws, row, yr_col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        exp_annual_rows.append(row)
        row += 1

    row += 1
    exp_ann_total_row = row
    set_cell(ws, row, "B", "Total Expansion Annual", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        refs = "+".join(f"{yr_col}{r}" for r in exp_annual_rows)
        set_cell(ws, row, yr_col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    # Store key rows
    ws._exp_total_row = exp_total_row
    ws._exp_ann_total_row = exp_ann_total_row
    ws._exp_annual_rows = exp_annual_rows
    ws._cum_golive_row = cum_golive_row
    ws._golive_zones_row = golive_zones_row
    ws._consult_impl_row = consult_impl_row
    ws._sub_row = sub_row
    ws._util_total_row = util_total_row

    return ws


# ============================================================
# SHEET 4: COMBINED SUMMARY
# ============================================================
def build_combined(wb, anchor_ws, panchor_ws):
    ws = wb.create_sheet("Combined Summary")
    ws.sheet_properties.tabColor = "660066"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    set_cell(ws, 1, "B", "Revenue Summary", font=Font(name=FONT_NAME, bold=True, size=14))

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, 3, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # ---- ANNUAL REVENUE ----
    set_cell(ws, 4, "B", "Annual Revenue", font=BOLD_FONT)

    # Row 5: Anchor Revenue (GAAP)
    set_cell(ws, 5, "B", "  Anchor Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 5, yr_col,
                 f"='Anchor Revenue'!{yr_col}{anchor_ws._ann_total_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    # Row 6: Non-GAAP Revenue
    set_cell(ws, 6, "B", "  Non-GAAP Revenue",
             font=Font(name=FONT_NAME, color="008000", italic=True))
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 6, yr_col,
                 f"='Anchor Revenue'!{yr_col}{anchor_ws._annual_rows[1]}"
                 f"+'Anchor Revenue'!{yr_col}{anchor_ws._annual_rows[2]}",
                 font=Font(name=FONT_NAME, color="008000", italic=True), fmt=CUR_FMT)

    # Row 7: Expansion Revenue
    set_cell(ws, 7, "B", "  Expansion Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 7, yr_col,
                 f"='Expansion Pipeline'!{yr_col}{panchor_ws._exp_ann_total_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    # Row 9: Total Anchor - incl. Non-GAAP
    set_cell(ws, 9, "B", "Total Anchor - incl. Non-GAAP", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 9, yr_col, f"={yr_col}5+{yr_col}6", font=BOLD_FONT, fmt=CUR_FMT)

    # Row 10: Total Revenue (GAAP)
    set_cell(ws, 10, "B", "Total Revenue", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 10, yr_col, f"={yr_col}5+{yr_col}7", font=BOLD_FONT, fmt=CUR_FMT)

    # Row 11: Total Revenue incl. Non-GAAP
    set_cell(ws, 11, "B", "Total Revenue incl. Non-GAAP", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 11, yr_col, f"={yr_col}9+{yr_col}7", font=BOLD_FONT, fmt=CUR_FMT)

    # Row 13: Cumulative Revenue
    set_cell(ws, 13, "B", "Cumulative Revenue", font=BOLD_FONT)
    set_cell(ws, 13, "C", "=C10", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, 13, yr_col, f"={prev_col}13+{yr_col}10", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- ZONE DEPLOYMENT TIMELINE ----
    set_cell(ws, 15, "B", "Zone Deployment Timeline", font=BOLD_FONT)

    # Row 17: Anchor Zones (cumulative) — use anchor cumulative zone helpers at Q4 of each year
    set_cell(ws, 17, "B", "  Anchor Zones (cumulative)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        q4_qi = 4 + 4 * yi
        q4_col = qcol(q4_qi)
        refs = "+".join(f"'Anchor Revenue'!{q4_col}{r}" for r in anchor_ws._cum_zone_rows)
        set_cell(ws, 17, yr_col, f"={refs}", font=BLACK_FONT, fmt=INT_FMT)

    # Row 18: Expansion Zones
    set_cell(ws, 18, "B", "  Expansion Zones (cumulative go-live)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        q4_qi = 4 + 4 * yi
        q4_col = qcol(q4_qi)
        set_cell(ws, 18, yr_col,
                 f"='Expansion Pipeline'!{q4_col}{panchor_ws._cum_golive_row}",
                 font=BLACK_FONT, fmt=INT_FMT)

    # Row 19: TOTAL DEPLOYED ZONES
    set_cell(ws, 19, "B", "Total Deployed Zones", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 19, yr_col, f"={yr_col}17+{yr_col}18", font=BOLD_FONT, fmt=INT_FMT)

    # Row 21: Revenue per Deployed Zone
    set_cell(ws, 21, "B", "Revenue per Deployed Zone", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 21, yr_col,
                 f'=IF({yr_col}19=0,0,{yr_col}10/{yr_col}19)',
                 font=BLACK_FONT, fmt=CUR_FMT)

    # ---- REVENUE BREAKDOWNS ----
    set_cell(ws, 23, "B", "Anchor Revenue Breakdown", font=BOLD_FONT)

    # annual_rows = [impl, nongaap_impl, nongaap_sub, lic, proc, mon, img]
    anchor_breakdown = [
        ("  Implementation", anchor_ws._annual_rows[0]),
        ("  Non-GAAP Impl + Consulting", anchor_ws._annual_rows[1]),
        ("  Non-GAAP Subscription", anchor_ws._annual_rows[2]),
        ("  Licensing", anchor_ws._annual_rows[3]),
        ("  Operations", anchor_ws._annual_rows[4]),
        ("  Maintenance", anchor_ws._annual_rows[5]),
        ("  Inspection", anchor_ws._annual_rows[6]),
    ]
    r = 24
    for label, src_row in anchor_breakdown:
        set_cell(ws, r, "B", label, font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            set_cell(ws, r, yr_col,
                     f"='Anchor Revenue'!{yr_col}{src_row}",
                     font=BLACK_FONT, fmt=CUR_FMT)
        r += 1

    r += 1
    set_cell(ws, r, "B", "EXPANSION BREAKDOWN", font=BOLD_FONT)
    r += 1

    panchor_breakdown = [
        ("  Consulting + Implementation", panchor_ws._exp_annual_rows[0]),
        ("  Subscription", panchor_ws._exp_annual_rows[1]),
        ("  Utilization", panchor_ws._exp_annual_rows[2]),
    ]
    for label, src_row in panchor_breakdown:
        set_cell(ws, r, "B", label, font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            set_cell(ws, r, yr_col,
                     f"='Expansion Pipeline'!{yr_col}{src_row}",
                     font=BLACK_FONT, fmt=CUR_FMT)
        r += 1

    # Store key rows for P&L and DCF cross-refs
    ws._anchor_rev_row = 5
    ws._nongaap_rev_row = 6
    ws._exp_rev_row = 7
    ws._total_anchor_incl_nongaap_row = 9
    ws._total_rev_row = 10
    ws._total_rev_incl_nongaap_row = 11

    return ws


# ============================================================
# SHEET 5: COST FORECAST
# ============================================================
def build_cost_forecast(wb, anchor_ws, panchor_ws):
    ws = wb.create_sheet("Cost Forecast")
    ws.sheet_properties.tabColor = "CC0000"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 23):
        ws.column_dimensions[get_column_letter(c)].width = 14

    set_cell(ws, 1, "B", "Cost Forecast", font=Font(name=FONT_NAME, bold=True, size=14))
    write_quarter_headers(ws, 3, 3)
    write_quarter_indices(ws, 4, 3)

    # ---- SECTION 1: ANCHOR SITE COSTS ----
    row = 6
    set_cell(ws, row, "B", "Anchor Site Costs", font=BOLD_FONT)
    row += 1
    set_cell(ws, row, "B", "  Non-Recurring Costs", font=BOLD_FONT)
    row += 1

    # NRC rows: zone-aware — NRC hits each quarter zones go live
    nrc_rows = []
    for si, s in enumerate(SYSTEMS):
        set_cell(ws, row, "B", f"    {s['name']}", font=BLACK_FONT)
        ramp_row = A_ZONE_RAMP_START + si
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            if qi > 8:
                # No anchor zone go-lives after Q4'28
                set_cell(ws, row, col, 0, font=BLACK_FONT, fmt=CUR_FMT)
            else:
                ramp_col = get_column_letter(qi + 2)  # qi=1->C, qi=8->J
                zones_ref = f"Assumptions!{ramp_col}${ramp_row}"
                # Check if any zones went live before this quarter
                if qi == 1:
                    cum_before = "0"
                else:
                    prev_end = get_column_letter(qi + 1)  # column before current
                    cum_before = f"SUM(Assumptions!$C${ramp_row}:{prev_end}${ramp_row})"
                formula = (
                    f"=IF({zones_ref}=0,0,"
                    f"IF({cum_before}=0,"
                    f"Assumptions!$C${A_NRC_FIRST}+({zones_ref}-1)*Assumptions!$C${A_NRC_ADDL},"
                    f"{zones_ref}*Assumptions!$C${A_NRC_ADDL}))"
                )
                set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        nrc_rows.append(row)
        row += 1

    row += 1
    anchor_nrc_total_row = row
    set_cell(ws, row, "B", "  Total Anchor NRC", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in nrc_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 2
    set_cell(ws, row, "B", "  Recurring Costs", font=BOLD_FONT)
    row += 1

    # Recurring rows: use cumulative zones from Anchor Revenue helpers
    rec_rows = []
    for si, s in enumerate(SYSTEMS):
        set_cell(ws, row, "B", f"    {s['name']}", font=BLACK_FONT)
        cum_zone_r = anchor_ws._cum_zone_rows[si]
        for qi in range(1, NUM_Q + 1):
            col = qcol(qi)
            cum_ref = f"'Anchor Revenue'!{col}{cum_zone_r}"
            formula = (
                f"=IF({cum_ref}=0,0,"
                f"Assumptions!$C${A_QRC}+({cum_ref}-1)*Assumptions!$C${A_QRC_ADDL})"
            )
            set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        rec_rows.append(row)
        row += 1

    row += 1
    anchor_rec_total_row = row
    set_cell(ws, row, "B", "  Total Anchor Recurring", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        refs = "+".join(f"{col}{r}" for r in rec_rows)
        set_cell(ws, row, col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 2
    anchor_cost_total_row = row
    set_cell(ws, row, "B", "Total Anchor Costs", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{anchor_nrc_total_row}+{col}{anchor_rec_total_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 2: EXPANSION SITE COSTS ----
    row += 2
    set_cell(ws, row, "B", "EXPANSION SITE COSTS", font=BOLD_FONT)
    row += 1
    set_cell(ws, row, "B", "  Non-Recurring Costs", font=BOLD_FONT)
    row += 1

    exp_nrc_row = row
    set_cell(ws, row, "B", "    NRC at Go-Live", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        formula = (
            f"='Expansion Pipeline'!{col}{panchor_ws._golive_zones_row}"
            f"*Assumptions!$C${A_BLEND_NRC}"
        )
        set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)

    row += 2
    exp_nrc_total_row = row
    set_cell(ws, row, "B", "  Total Expansion NRC", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{exp_nrc_row}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 2
    set_cell(ws, row, "B", "  Recurring Costs", font=BOLD_FONT)
    row += 1

    exp_rec_row = row
    set_cell(ws, row, "B", "    Recurring from Go-Live Zones", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        formula = (
            f"='Expansion Pipeline'!{col}{panchor_ws._cum_golive_row}"
            f"*Assumptions!$C${A_BLEND_QRC}"
        )
        set_cell(ws, row, col, formula, font=BLACK_FONT, fmt=CUR_FMT)

    row += 2
    exp_rec_total_row = row
    set_cell(ws, row, "B", "  Total Expansion Recurring", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{exp_rec_row}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 2
    exp_cost_total_row = row
    set_cell(ws, row, "B", "Total Expansion Costs", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{exp_nrc_total_row}+{col}{exp_rec_total_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # ---- SECTION 3: TOTAL COSTS ----
    row += 2
    set_cell(ws, row, "B", "Total Costs", font=BOLD_FONT)
    row += 1

    total_nrc_row = row
    set_cell(ws, row, "B", "  Total NRC", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{anchor_nrc_total_row}+{col}{exp_nrc_total_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)
    row += 1

    total_rec_row = row
    set_cell(ws, row, "B", "  Total Recurring", font=BLACK_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{anchor_rec_total_row}+{col}{exp_rec_total_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    row += 2
    grand_total_row = row
    set_cell(ws, row, "B", "Total Cost", font=BOLD_FONT)
    for qi in range(1, NUM_Q + 1):
        col = qcol(qi)
        set_cell(ws, row, col, f"={col}{total_nrc_row}+{col}{total_rec_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # ---- ANNUAL SUMMARY ----
    row += 2
    set_cell(ws, row, "B", "Annual Summary", font=BOLD_FONT)
    row += 1

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, row, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))
    row += 1

    ann_src = [
        ("Anchor NRC", anchor_nrc_total_row),
        ("Anchor Recurring", anchor_rec_total_row),
        ("Expansion NRC", exp_nrc_total_row),
        ("Expansion Recurring", exp_rec_total_row),
    ]
    cost_annual_rows = []
    for label, src_row in ann_src:
        set_cell(ws, row, "B", f"  {label}", font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            q_start = 1 + 4 * yi
            q_cols = [qcol(q_start + q) for q in range(4)]
            formula = "=" + "+".join(f"'{ws.title}'!{c}{src_row}" for c in q_cols)
            set_cell(ws, row, yr_col, formula, font=BLACK_FONT, fmt=CUR_FMT)
        cost_annual_rows.append(row)
        row += 1

    row += 1
    ann_total_cost_row = row
    set_cell(ws, row, "B", "Total Annual Cost", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        refs = "+".join(f"{yr_col}{r}" for r in cost_annual_rows)
        set_cell(ws, row, yr_col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    row += 1
    set_cell(ws, row, "B", "Cumulative Cost", font=BOLD_FONT)
    set_cell(ws, row, "C", f"=C{ann_total_cost_row}", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, row, yr_col, f"={prev_col}{row}+{yr_col}{ann_total_cost_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    # Store key rows for P&L cross-refs
    ws._ann_total_cost_row = ann_total_cost_row
    ws._cost_annual_rows = cost_annual_rows  # [Anchor NRC, Anchor Rec, Expansion NRC, Expansion Rec]

    return ws


# ============================================================
# SHEET 6: P&L SUMMARY
# ============================================================
def build_pnl(wb, anchor_ws, panchor_ws, cost_ws, combined_ws):
    ws = wb.create_sheet("P&L Summary")
    ws.sheet_properties.tabColor = "006600"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    set_cell(ws, 1, "B", "P&L Summary", font=Font(name=FONT_NAME, bold=True, size=14))

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, 3, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # ---- REVENUE ----
    set_cell(ws, 5, "B", "Revenue", font=BOLD_FONT)

    set_cell(ws, 6, "B", "  Anchor Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 6, yr_col,
                 f"='Combined Summary'!{yr_col}{combined_ws._anchor_rev_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 7, "B", "  Non-GAAP Revenue", font=Font(name=FONT_NAME, color="008000", italic=True))
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 7, yr_col,
                 f"='Combined Summary'!{yr_col}{combined_ws._nongaap_rev_row}",
                 font=Font(name=FONT_NAME, color="008000", italic=True), fmt=CUR_FMT)

    set_cell(ws, 8, "B", "  Expansion Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 8, yr_col,
                 f"='Combined Summary'!{yr_col}{combined_ws._exp_rev_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 10, "B", "Total Revenue (GAAP)", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 10, yr_col, f"={yr_col}6+{yr_col}8", font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 11, "B", "Total Revenue incl. Non-GAAP", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 11, yr_col, f"={yr_col}6+{yr_col}7+{yr_col}8", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- COST OF GOODS SOLD (COGS) ----
    set_cell(ws, 13, "B", "Cost of Goods Sold", font=BOLD_FONT)

    anchor_rec_ann = cost_ws._cost_annual_rows[1]   # Anchor Recurring
    panchor_rec_ann = cost_ws._cost_annual_rows[3]   # Expansion Recurring

    set_cell(ws, 14, "B", "  Anchor Recurring Costs", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 14, yr_col,
                 f"='Cost Forecast'!{yr_col}{anchor_rec_ann}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 15, "B", "  Expansion Recurring Costs", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 15, yr_col,
                 f"='Cost Forecast'!{yr_col}{panchor_rec_ann}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 17, "B", "Total COGS", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 17, yr_col, f"={yr_col}14+{yr_col}15", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- GROSS PROFIT ----
    set_cell(ws, 19, "B", "Gross Profit", font=BOLD_FONT)

    set_cell(ws, 20, "B", "Gross Profit (GAAP)", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 20, yr_col, f"={yr_col}10-{yr_col}17", font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 21, "B", "  Gross Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 21, yr_col, f"=IF({yr_col}10=0,0,{yr_col}20/{yr_col}10)",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # ---- DEPRECIATION (CapEx Amortization) ----
    set_cell(ws, 23, "B", "Depreciation (CapEx Amortization)", font=BOLD_FONT)

    anchor_nrc_ann = cost_ws._cost_annual_rows[0]    # Anchor NRC
    panchor_nrc_ann = cost_ws._cost_annual_rows[2]   # Expansion NRC

    set_cell(ws, 24, "B", "  Anchor NRC Amortization", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        terms = []
        for j in range(yi + 1):
            j_col = get_column_letter(3 + j)
            terms.append(
                f"IF({yi - j}<Assumptions!$C${A_AMORT_YRS},"
                f"'Cost Forecast'!{j_col}{anchor_nrc_ann}/Assumptions!$C${A_AMORT_YRS},0)"
            )
        set_cell(ws, 24, yr_col, f"={'+'.join(terms)}", font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 25, "B", "  Expansion NRC Amortization", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        terms = []
        for j in range(yi + 1):
            j_col = get_column_letter(3 + j)
            terms.append(
                f"IF({yi - j}<Assumptions!$C${A_AMORT_YRS},"
                f"'Cost Forecast'!{j_col}{panchor_nrc_ann}/Assumptions!$C${A_AMORT_YRS},0)"
            )
        set_cell(ws, 25, yr_col, f"={'+'.join(terms)}", font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 26, "B", "Total Depreciation", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 26, yr_col, f"={yr_col}24+{yr_col}25", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- OPERATING EXPENSES ----
    set_cell(ws, 28, "B", "Operating Expenses", font=BOLD_FONT)

    set_cell(ws, 29, "B", "  G&A OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 29, yr_col, f"={yr_col}10*Assumptions!$C${A_GA_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 30, "B", "  R&D OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 30, yr_col, f"={yr_col}10*Assumptions!$C${A_RD_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 31, "B", "Total OpEx", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 31, yr_col, f"={yr_col}29+{yr_col}30", font=BOLD_FONT, fmt=CUR_FMT)

    # ---- EBITDA ----
    set_cell(ws, 33, "B", "EBITDA", font=BOLD_FONT)

    set_cell(ws, 34, "B", "EBITDA (GAAP)", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 34, yr_col, f"={yr_col}20-{yr_col}26-{yr_col}31",
                 font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 35, "B", "  EBITDA Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 35, yr_col, f"=IF({yr_col}10=0,0,{yr_col}34/{yr_col}10)",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # ---- CUMULATIVE ----
    set_cell(ws, 37, "B", "Cumulative", font=BOLD_FONT)

    set_cell(ws, 38, "B", "Cumulative Revenue", font=BOLD_FONT)
    set_cell(ws, 38, "C", "=C10", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, 38, yr_col, f"={prev_col}38+{yr_col}10", font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 39, "B", "Cumulative COGS", font=BOLD_FONT)
    set_cell(ws, 39, "C", "=C17", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, 39, yr_col, f"={prev_col}39+{yr_col}17", font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 40, "B", "Cumulative EBITDA", font=BOLD_FONT)
    set_cell(ws, 40, "C", "=C34", font=BOLD_FONT, fmt=CUR_FMT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, 40, yr_col, f"={prev_col}40+{yr_col}34", font=BOLD_FONT, fmt=CUR_FMT)

    return ws


# ============================================================
# SHEET 7: VALUATION - ANCHOR
# ============================================================
def build_valuation_anchor(wb, anchor_ws, combined_ws, cost_ws):
    ws = wb.create_sheet("Valuation - Anchor")
    ws.sheet_properties.tabColor = "996633"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    set_cell(ws, 1, "B", "Valuation \u2014 Anchor (Non-GAAP Basis)",
             font=Font(name=FONT_NAME, bold=True, size=14))

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, 3, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # Key assumptions
    set_cell(ws, 5, "B", "Key Assumptions", font=BOLD_FONT)
    set_cell(ws, 6, "B", "Discount Rate")
    set_cell(ws, 6, "C", 0.25, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, 7, "B", "Terminal Revenue Multiple")
    set_cell(ws, 7, "C", 10, font=BLUE_FONT, fmt=MULT_FMT, fill=YELLOW_FILL)

    # Revenue
    set_cell(ws, 9, "B", "Revenue", font=BOLD_FONT)

    # Pull from Anchor annual rows: [impl, nongaap_impl, nongaap_sub, lic, proc, mon, img]
    rev_labels = [
        ("  Implementation", anchor_ws._annual_rows[0]),
        ("  Non-GAAP Impl + Consulting", anchor_ws._annual_rows[1]),
        ("  Non-GAAP Subscription", anchor_ws._annual_rows[2]),
        ("  Licensing", anchor_ws._annual_rows[3]),
        ("  Operations", anchor_ws._annual_rows[4]),
        ("  Maintenance", anchor_ws._annual_rows[5]),
        ("  Inspection", anchor_ws._annual_rows[6]),
    ]
    rev_detail_rows = []
    r = 10
    for label, src_row in rev_labels:
        set_cell(ws, r, "B", label, font=BLACK_FONT)
        for yi in range(5):
            yr_col = get_column_letter(3 + yi)
            set_cell(ws, r, yr_col,
                     f"='Anchor Revenue'!{yr_col}{src_row}",
                     font=BLACK_FONT, fmt=CUR_FMT)
        rev_detail_rows.append(r)
        r += 1

    r += 1
    rev_total_row = r
    set_cell(ws, r, "B", "Total Anchor Revenue (Non-GAAP)", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        refs = "+".join(f"{yr_col}{dr}" for dr in rev_detail_rows)
        set_cell(ws, r, yr_col, f"={refs}", font=BOLD_FONT, fmt=CUR_FMT)

    r += 1
    set_cell(ws, r, "B", "  YoY Growth", font=BLACK_FONT)
    growth_row = r
    set_cell(ws, r, "C", "", font=BLACK_FONT)  # no growth for year 1
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, r, yr_col,
                 f"=IF({prev_col}{rev_total_row}=0,0,"
                 f"({yr_col}{rev_total_row}-{prev_col}{rev_total_row})/{prev_col}{rev_total_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # DCF analysis
    r += 2
    set_cell(ws, r, "B", "DCF Analysis", font=BOLD_FONT)
    r += 1
    year_num_row = r
    set_cell(ws, r, "B", "Year Number", font=BLACK_FONT)
    for yi in range(5):
        set_cell(ws, r, get_column_letter(3 + yi), yi + 1, font=BLACK_FONT, fmt=INT_FMT)

    r += 1
    df_row = r
    set_cell(ws, r, "B", "Discount Factor", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"=1/(1+$C$6)^{yr_col}{year_num_row}",
                 font=BLACK_FONT, fmt='0.0000')

    r += 1
    pv_rev_row = r
    set_cell(ws, r, "B", "PV of Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{rev_total_row}*{yr_col}{df_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 2
    tv_row = r
    set_cell(ws, r, "B", "Terminal Value (Year 5)", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{rev_total_row}*$C$7", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    pv_tv_row = r
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{tv_row}*G{df_row}", font=BLACK_FONT, fmt=CUR_FMT)

    # Valuation summary
    r += 2
    set_cell(ws, r, "B", "Valuation Summary", font=BOLD_FONT)
    r += 1
    set_cell(ws, r, "B", "Sum of Discounted Revenue", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=SUM(C{pv_rev_row}:G{pv_rev_row})", font=BLACK_FONT, fmt=CUR_FMT)
    sum_pv_row = r

    r += 1
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=G{pv_tv_row}", font=BLACK_FONT, fmt=CUR_FMT)
    pv_tv_summary_row = r

    r += 1
    set_cell(ws, r, "B", "Enterprise Value (Revenue)", font=Font(name=FONT_NAME, bold=True, size=12))
    set_cell(ws, r, "C", f"=C{sum_pv_row}+C{pv_tv_summary_row}",
             font=Font(name=FONT_NAME, bold=True, size=12), fmt=CUR_FMT)
    rev_ev_row = r

    # ================================================================
    # EBITDA-BASED VALUATION (appended below revenue DCF)
    # ================================================================
    r += 3
    set_cell(ws, r, "B", "EBITDA-BASED VALUATION",
             font=Font(name=FONT_NAME, bold=True, size=14))

    # Cost & Margin Analysis
    r += 2
    set_cell(ws, r, "B", "Cost & Margin Analysis", font=BOLD_FONT)

    r += 1
    eb_rev_row = r
    set_cell(ws, r, "B", "  Total Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{rev_total_row}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_cogs_row = r
    anchor_rec_ann = cost_ws._cost_annual_rows[1]
    set_cell(ws, r, "B", "  COGS (Anchor Recurring)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"='Cost Forecast'!{yr_col}{anchor_rec_ann}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_gp_row = r
    set_cell(ws, r, "B", "Gross Profit", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}-{yr_col}{eb_cogs_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    r += 1
    set_cell(ws, r, "B", "  Gross Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"=IF({yr_col}{eb_rev_row}=0,0,{yr_col}{eb_gp_row}/{yr_col}{eb_rev_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    r += 1
    eb_depr_row = r
    anchor_nrc_ann = cost_ws._cost_annual_rows[0]
    set_cell(ws, r, "B", "  Depreciation (NRC Amort.)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        terms = []
        for j in range(yi + 1):
            j_col = get_column_letter(3 + j)
            terms.append(
                f"IF({yi - j}<Assumptions!$C${A_AMORT_YRS},"
                f"'Cost Forecast'!{j_col}{anchor_nrc_ann}/Assumptions!$C${A_AMORT_YRS},0)"
            )
        set_cell(ws, r, yr_col, f"={'+'.join(terms)}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_ga_row = r
    set_cell(ws, r, "B", "  G&A OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}*Assumptions!$C${A_GA_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_rd_row = r
    set_cell(ws, r, "B", "  R&D OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}*Assumptions!$C${A_RD_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_ebitda_row = r
    set_cell(ws, r, "B", "EBITDA", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"={yr_col}{eb_gp_row}-{yr_col}{eb_depr_row}-{yr_col}{eb_ga_row}-{yr_col}{eb_rd_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    r += 1
    set_cell(ws, r, "B", "  EBITDA Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"=IF({yr_col}{eb_rev_row}=0,0,{yr_col}{eb_ebitda_row}/{yr_col}{eb_rev_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # EBITDA-Based DCF
    r += 2
    set_cell(ws, r, "B", "EBITDA-BASED DCF", font=BOLD_FONT)

    r += 1
    eb_yr_row = r
    set_cell(ws, r, "B", "Year Number", font=BLACK_FONT)
    for yi in range(5):
        set_cell(ws, r, get_column_letter(3 + yi), yi + 1, font=BLACK_FONT, fmt=INT_FMT)

    r += 1
    eb_df_row = r
    set_cell(ws, r, "B", "Discount Factor", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"=1/(1+$C$6)^{yr_col}{eb_yr_row}",
                 font=BLACK_FONT, fmt='0.0000')

    r += 1
    eb_pv_row = r
    set_cell(ws, r, "B", "PV of EBITDA", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_ebitda_row}*{yr_col}{eb_df_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 2
    eb_tv_row = r
    set_cell(ws, r, "B", "Terminal Value (Year 5 EBITDA)", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{eb_ebitda_row}*Assumptions!$C${A_EBITDA_MULT}",
             font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_pv_tv_row = r
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{eb_tv_row}*G{eb_df_row}", font=BLACK_FONT, fmt=CUR_FMT)

    # EBITDA Valuation Summary
    r += 2
    set_cell(ws, r, "B", "EBITDA Valuation Summary", font=BOLD_FONT)

    r += 1
    eb_sum_pv = r
    set_cell(ws, r, "B", "Sum of Discounted EBITDA", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=SUM(C{eb_pv_row}:G{eb_pv_row})", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_pv_tv_summ = r
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=G{eb_pv_tv_row}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    ebitda_ev_row = r
    set_cell(ws, r, "B", "Enterprise Value (EBITDA)", font=Font(name=FONT_NAME, bold=True, size=12))
    set_cell(ws, r, "C", f"=C{eb_sum_pv}+C{eb_pv_tv_summ}",
             font=Font(name=FONT_NAME, bold=True, size=12), fmt=CUR_FMT)

    # Comparison
    r += 2
    set_cell(ws, r, "B", "Comparison", font=BOLD_FONT)
    r += 1
    set_cell(ws, r, "B", "  Revenue-Based EV", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=C{rev_ev_row}", font=BLACK_FONT, fmt=CUR_FMT)
    r += 1
    set_cell(ws, r, "B", "  EBITDA-Based EV", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=C{ebitda_ev_row}", font=BLACK_FONT, fmt=CUR_FMT)

    return ws


# ============================================================
# SHEET 8: VALUATION - TOTAL
# ============================================================
def build_valuation_total(wb, combined_ws, cost_ws):
    ws = wb.create_sheet("Valuation - Total")
    ws.sheet_properties.tabColor = "996633"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    for c in range(3, 8):
        ws.column_dimensions[get_column_letter(c)].width = 16

    set_cell(ws, 1, "B", "Valuation \u2014 Total Company",
             font=Font(name=FONT_NAME, bold=True, size=14))

    years = [2027, 2028, 2029, 2030, 2031]
    for i, yr in enumerate(years):
        set_cell(ws, 3, get_column_letter(3 + i), str(yr), font=WHITE_BOLD, fill=DARK_BLUE_FILL,
                 alignment=Alignment(horizontal="center"))

    # Key assumptions
    set_cell(ws, 5, "B", "Key Assumptions", font=BOLD_FONT)
    set_cell(ws, 6, "B", "Discount Rate")
    set_cell(ws, 6, "C", 0.25, font=BLUE_FONT, fmt=PCT_FMT, fill=YELLOW_FILL)
    set_cell(ws, 7, "B", "Terminal Revenue Multiple")
    set_cell(ws, 7, "C", 10, font=BLUE_FONT, fmt=MULT_FMT, fill=YELLOW_FILL)

    # Revenue
    set_cell(ws, 9, "B", "Revenue", font=BOLD_FONT)

    set_cell(ws, 10, "B", "  Anchor Revenue (GAAP)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 10, yr_col,
                 f"='Combined Summary'!{yr_col}{combined_ws._anchor_rev_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    set_cell(ws, 11, "B", "  Expansion Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, 11, yr_col,
                 f"='Combined Summary'!{yr_col}{combined_ws._exp_rev_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    rev_total_row = 13
    set_cell(ws, rev_total_row, "B", "Total Revenue", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, rev_total_row, yr_col, f"={yr_col}10+{yr_col}11",
                 font=BOLD_FONT, fmt=CUR_FMT)

    set_cell(ws, 14, "B", "  YoY Growth", font=BLACK_FONT)
    for yi in range(1, 5):
        yr_col = get_column_letter(3 + yi)
        prev_col = get_column_letter(2 + yi)
        set_cell(ws, 14, yr_col,
                 f"=IF({prev_col}{rev_total_row}=0,0,"
                 f"({yr_col}{rev_total_row}-{prev_col}{rev_total_row})/{prev_col}{rev_total_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # DCF analysis
    set_cell(ws, 16, "B", "DCF Analysis", font=BOLD_FONT)

    year_num_row = 17
    set_cell(ws, year_num_row, "B", "Year Number", font=BLACK_FONT)
    for yi in range(5):
        set_cell(ws, year_num_row, get_column_letter(3 + yi), yi + 1, font=BLACK_FONT, fmt=INT_FMT)

    df_row = 18
    set_cell(ws, df_row, "B", "Discount Factor", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, df_row, yr_col, f"=1/(1+$C$6)^{yr_col}{year_num_row}",
                 font=BLACK_FONT, fmt='0.0000')

    pv_rev_row = 19
    set_cell(ws, pv_rev_row, "B", "PV of Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, pv_rev_row, yr_col, f"={yr_col}{rev_total_row}*{yr_col}{df_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    tv_row = 21
    set_cell(ws, tv_row, "B", "Terminal Value (Year 5)", font=BLACK_FONT)
    set_cell(ws, tv_row, "G", f"=G{rev_total_row}*$C$7", font=BLACK_FONT, fmt=CUR_FMT)

    pv_tv_row = 22
    set_cell(ws, pv_tv_row, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, pv_tv_row, "G", f"=G{tv_row}*G{df_row}", font=BLACK_FONT, fmt=CUR_FMT)

    # Valuation summary
    set_cell(ws, 24, "B", "Valuation Summary", font=BOLD_FONT)

    sum_pv_row = 25
    set_cell(ws, sum_pv_row, "B", "Sum of Discounted Revenue", font=BLACK_FONT)
    set_cell(ws, sum_pv_row, "C", f"=SUM(C{pv_rev_row}:G{pv_rev_row})",
             font=BLACK_FONT, fmt=CUR_FMT)

    pv_tv_summary_row = 26
    set_cell(ws, pv_tv_summary_row, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, pv_tv_summary_row, "C", f"=G{pv_tv_row}", font=BLACK_FONT, fmt=CUR_FMT)

    ev_row = 27
    set_cell(ws, ev_row, "B", "Enterprise Value (Revenue)", font=Font(name=FONT_NAME, bold=True, size=12))
    set_cell(ws, ev_row, "C", f"=C{sum_pv_row}+C{pv_tv_summary_row}",
             font=Font(name=FONT_NAME, bold=True, size=12), fmt=CUR_FMT)
    rev_ev_row = ev_row

    # ================================================================
    # EBITDA-BASED VALUATION (appended below revenue DCF)
    # ================================================================
    r = 30
    set_cell(ws, r, "B", "EBITDA-BASED VALUATION",
             font=Font(name=FONT_NAME, bold=True, size=14))

    # Cost & Margin Analysis
    r += 2
    set_cell(ws, r, "B", "Cost & Margin Analysis", font=BOLD_FONT)

    r += 1
    eb_rev_row = r
    set_cell(ws, r, "B", "  Total Revenue", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{rev_total_row}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_cogs_row = r
    anchor_rec_ann = cost_ws._cost_annual_rows[1]
    panchor_rec_ann = cost_ws._cost_annual_rows[3]
    set_cell(ws, r, "B", "  COGS (All Recurring)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"='Cost Forecast'!{yr_col}{anchor_rec_ann}"
                 f"+'Cost Forecast'!{yr_col}{panchor_rec_ann}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_gp_row = r
    set_cell(ws, r, "B", "Gross Profit", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}-{yr_col}{eb_cogs_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    r += 1
    set_cell(ws, r, "B", "  Gross Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"=IF({yr_col}{eb_rev_row}=0,0,{yr_col}{eb_gp_row}/{yr_col}{eb_rev_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    r += 1
    eb_depr_row = r
    anchor_nrc_ann = cost_ws._cost_annual_rows[0]
    panchor_nrc_ann = cost_ws._cost_annual_rows[2]
    set_cell(ws, r, "B", "  Depreciation (NRC Amort.)", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        terms = []
        for j in range(yi + 1):
            j_col = get_column_letter(3 + j)
            terms.append(
                f"IF({yi - j}<Assumptions!$C${A_AMORT_YRS},"
                f"('Cost Forecast'!{j_col}{anchor_nrc_ann}+'Cost Forecast'!{j_col}{panchor_nrc_ann})"
                f"/Assumptions!$C${A_AMORT_YRS},0)"
            )
        set_cell(ws, r, yr_col, f"={'+'.join(terms)}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_ga_row = r
    set_cell(ws, r, "B", "  G&A OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}*Assumptions!$C${A_GA_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_rd_row = r
    set_cell(ws, r, "B", "  R&D OpEx", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_rev_row}*Assumptions!$C${A_RD_PCT}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_ebitda_row = r
    set_cell(ws, r, "B", "EBITDA", font=BOLD_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"={yr_col}{eb_gp_row}-{yr_col}{eb_depr_row}-{yr_col}{eb_ga_row}-{yr_col}{eb_rd_row}",
                 font=BOLD_FONT, fmt=CUR_FMT)

    r += 1
    set_cell(ws, r, "B", "  EBITDA Margin %", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col,
                 f"=IF({yr_col}{eb_rev_row}=0,0,{yr_col}{eb_ebitda_row}/{yr_col}{eb_rev_row})",
                 font=BLACK_FONT, fmt=PCT_FMT)

    # EBITDA-Based DCF
    r += 2
    set_cell(ws, r, "B", "EBITDA-BASED DCF", font=BOLD_FONT)

    r += 1
    eb_yr_row = r
    set_cell(ws, r, "B", "Year Number", font=BLACK_FONT)
    for yi in range(5):
        set_cell(ws, r, get_column_letter(3 + yi), yi + 1, font=BLACK_FONT, fmt=INT_FMT)

    r += 1
    eb_df_row = r
    set_cell(ws, r, "B", "Discount Factor", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"=1/(1+$C$6)^{yr_col}{eb_yr_row}",
                 font=BLACK_FONT, fmt='0.0000')

    r += 1
    eb_pv_row = r
    set_cell(ws, r, "B", "PV of EBITDA", font=BLACK_FONT)
    for yi in range(5):
        yr_col = get_column_letter(3 + yi)
        set_cell(ws, r, yr_col, f"={yr_col}{eb_ebitda_row}*{yr_col}{eb_df_row}",
                 font=BLACK_FONT, fmt=CUR_FMT)

    r += 2
    eb_tv_row = r
    set_cell(ws, r, "B", "Terminal Value (Year 5 EBITDA)", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{eb_ebitda_row}*Assumptions!$C${A_EBITDA_MULT}",
             font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_pv_tv_row = r
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "G", f"=G{eb_tv_row}*G{eb_df_row}", font=BLACK_FONT, fmt=CUR_FMT)

    # EBITDA Valuation Summary
    r += 2
    set_cell(ws, r, "B", "EBITDA Valuation Summary", font=BOLD_FONT)

    r += 1
    eb_sum_pv = r
    set_cell(ws, r, "B", "Sum of Discounted EBITDA", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=SUM(C{eb_pv_row}:G{eb_pv_row})", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    eb_pv_tv_summ = r
    set_cell(ws, r, "B", "PV of Terminal Value", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=G{eb_pv_tv_row}", font=BLACK_FONT, fmt=CUR_FMT)

    r += 1
    ebitda_ev_row = r
    set_cell(ws, r, "B", "Enterprise Value (EBITDA)", font=Font(name=FONT_NAME, bold=True, size=12))
    set_cell(ws, r, "C", f"=C{eb_sum_pv}+C{eb_pv_tv_summ}",
             font=Font(name=FONT_NAME, bold=True, size=12), fmt=CUR_FMT)

    # Comparison
    r += 2
    set_cell(ws, r, "B", "Comparison", font=BOLD_FONT)
    r += 1
    set_cell(ws, r, "B", "  Revenue-Based EV", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=C{rev_ev_row}", font=BLACK_FONT, fmt=CUR_FMT)
    r += 1
    set_cell(ws, r, "B", "  EBITDA-Based EV", font=BLACK_FONT)
    set_cell(ws, r, "C", f"=C{ebitda_ev_row}", font=BLACK_FONT, fmt=CUR_FMT)

    return ws


# ============================================================
# SHEET 9: SOURCES (trimmed reference)
# ============================================================
def build_sources(wb):
    ws = wb.create_sheet("Sources")
    ws.sheet_properties.tabColor = "336699"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80

    set_cell(ws, 1, "B", "Sources & References", font=Font(name=FONT_NAME, bold=True, size=14))

    # Curated parameter → source pairs, grouped by section
    sections = [
        ("Use Case Assumptions", [
            ("Autonomous Picking \u2014 Avg Duration", "Warehouse robotics cycle time benchmarks; Locus Robotics case studies"),
            ("Autonomous Picking \u2014 Volume", "MHIA warehouse throughput studies; Berkshire Grey white paper"),
            ("Pallet Sorting \u2014 Avg Duration", "Automated sortation/inline system benchmarks"),
            ("Pallet Sorting \u2014 Volume", "WERC warehouse operations survey; DHL/Amazon case studies"),
            ("Inventory Scanning \u2014 Avg Duration", "Warehouse cycle count benchmarks; Zebra Technologies inventory study"),
            ("Inventory Scanning \u2014 Volume", "Systematic review robotic governance; industry training standards; Apex Fulfillment"),
            ("Zone Transfer \u2014 Avg Duration", "robotic transfer time benchmarks"),
            ("Zone Transfer \u2014 Volume", "logistics industry cost-effectiveness; pilot program data; logistics analyst reports"),
            ("Mixed-Case Depalletizing \u2014 Avg Duration", "Multi-specialty robotic case duration benchmarks"),
            ("Mixed-Case Depalletizing \u2014 Volume", "Atlas Logistics; Titan Distribution; Atlas Logistics; Fetch Robotics"),
        ]),
        ("Utilization Pricing", [
            ("Tier 1 \u2014 $/minute", "Arctura pricing framework; robotic operations economics analysis"),
            ("Tier 2 \u2014 $/minute", "Arctura pricing framework"),
            ("Maintenance \u2014 $/robot/month", "Arctura pricing; predictive maintenance industry benchmarks"),
            ("Inspection \u2014 $/scan", "Arctura pricing framework"),
        ]),
        ("Anchor Config", [
            ("Zone Count (Atlas)", "Atlas Logistics operational footprint; warehouse layout analysis"),
            ("Implementation MSRP", "Arctura cost model (site_cost_summary.xlsx)"),
            ("Annual Licensing Fee", "Arctura pricing framework"),
            ("Payment Terms (50/30/20)", "Arctura contract structure"),
            ("Implementation Timeline", "Arctura implementation playbook"),
        ]),
        ("Expansion Pricing", [
            ("Implementation Fee (per zone)", "Arctura expansion cost model"),
            ("Program Consulting Fee", "Arctura expansion pricing"),
            ("Annual Subscription", "Arctura expansion pricing"),
        ]),
        ("Ramp & Decay", [
            ("Hockey Stick Ramp (Q1\u2013Q8)", "Healthcare technology adoption curves; enterprise SaaS benchmarks"),
            ("Site Volume Decay", "Multi-zone model; operational complexity scaling"),
        ]),
        ("Maintenance & Inspection", [
            ("Robots per Zone", "Atlas Logistics warehouse operations paper; warehouse robotics density analysis"),
            ("Scans per Zone per Month", "Atlas Logistics quality inspection; industry benchmarks; OSHA warehouse guidelines"),
        ]),
        ("Site Costs", [
            ("Non-Recurring Cost (per site)", "Arctura site cost summary (site_cost_summary.xlsx)"),
            ("Monthly Recurring Cost", "Arctura site cost summary (site_cost_summary.xlsx)"),
            ("Additional Zone Discount (25%)", "Internal analysis"),
        ]),
        ("Atlas Logistics Baselines", [
            ("Annual Shipments (180K/yr)", "Atlas Logistics public operations data"),
            ("Warehouse Throughput (50K units/day)", "Atlas Logistics operations dashboard"),
            ("Annual Picks (>15M/yr)", "Atlas Logistics fulfillment metrics"),
            ("Current Robot Fleet (120)", "Atlas Logistics automation program summary"),
            ("QC Inspections (~400K/yr)", "Atlas Logistics quality program data"),
            ("Warehouse Sq Footage (2.5M)", "Atlas Logistics facility overview"),
        ]),
    ]

    row = 3
    for section_name, entries in sections:
        set_cell(ws, row, "B", section_name, font=BOLD_FONT)
        row += 1
        for param, source in entries:
            set_cell(ws, row, "B", param, font=BLACK_FONT)
            set_cell(ws, row, "C", source, font=Font(name=FONT_NAME, color="808080"))
            row += 1
        row += 1  # blank row between sections

    return ws


# ============================================================
# SHEET 10: SENSITIVITY ANALYSIS
# ============================================================
def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_properties.tabColor = "FF6600"

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10

    # Fonts for dot indicators
    dot_green = Font(name=FONT_NAME, bold=True, color="008000")   # ●●● major
    dot_orange = Font(name=FONT_NAME, bold=True, color="FF8C00")  # ●● moderate
    dot_gray = Font(name=FONT_NAME, bold=True, color="808080")    # ● minor
    dot_none = Font(name=FONT_NAME, color="808080")               # — none

    def dot_str(n):
        if n == 3: return "●●●"
        if n == 2: return "●●"
        if n == 1: return "●"
        return "—"

    def dot_font(n):
        if n == 3: return dot_green
        if n == 2: return dot_orange
        if n == 1: return dot_gray
        return dot_none

    # Sensitivity data: (tier, name, default, mechanism, rev_dots, ebitda_dots, ev_dots)
    SENSITIVITY_DATA = [
        # Tier 1 — Highest Impact (>$20M EV swing per ±20%)
        (1, "Expansion New Partners Signed", "43 total", "Scales all expansion revenue", 3, 3, 3),
        (1, "Avg Zones per Partner (Expansion)", "3", "Multiplier on all expansion zone-driven revenue", 3, 3, 3),
        (1, "Tier 1 $/minute", "$3.50", "Prices 4 of 5 operation types (largest revenue stream)", 3, 3, 3),
        (1, "Terminal Revenue Multiple", "10x", "Direct multiplier on terminal value (~80% of revenue EV)", 0, 0, 3),
        (1, "EBITDA Terminal Multiple", "12x", "Direct multiplier on EBITDA terminal value", 0, 0, 3),
        (1, "Discount Rate", "25%", "Exponential effect on all PV calculations", 0, 0, 3),
        (1, "Expansion Annual Subscription", "$95k/zone/yr", "Recurring revenue on cumulative zone-years", 3, 3, 3),

        # Tier 2 — High Impact ($5-20M EV swing per ±20%)
        (2, "Hockey Stick Ramp (Q1-Q8)", "12%→100%", "Multiplier on all utilization revenue timing", 3, 3, 3),
        (2, "Use Case: Ops/Wk (all 5)", "5-24/wk", "Volume driver for operations revenue", 3, 3, 3),
        (2, "Use Case: Avg Minutes (all 5)", "15-90 min", "Duration × price driver for operations", 3, 3, 3),
        (2, "Maintenance $/robot/month", "$40", "Prices maintenance stream", 2, 2, 2),
        (2, "Robots per Zone", "80", "Volume driver for maintenance revenue", 2, 2, 2),
        (2, "R&D OpEx %", "18%", "Largest OpEx line item", 0, 2, 2),
        (2, "G&A OpEx %", "12%", "Second OpEx line item", 0, 2, 2),
        (2, "Per-Zone Implementation MSRP", "$350k", "Drives anchor implementation revenue", 2, 0, 2),
        (2, "Per-Zone Annual License", "$85k", "Drives anchor licensing + Non-GAAP subscription", 2, 0, 2),
        (2, "Zone Counts (Anchor, 5 partners)", "6,4,5,3,2", "Scales all anchor per-zone revenue", 2, 2, 2),
        (2, "Expansion Impl Timeline", "8 months", "Delays go-live → delays utilization revenue", 2, 2, 2),

        # Tier 3 — Moderate Impact ($1-5M EV swing per ±20%)
        (3, "Site Volume Decay (Zone 1-6)", "100%→65%", "Reduces per-zone operations volumes", 2, 2, 0),
        (3, "Inspection $/scan", "$12", "Prices inspection stream", 1, 1, 1),
        (3, "Scans/Zone/Month", "250", "Volume driver for inspection", 1, 1, 1),
        (3, "Tier 2 $/minute", "$0.75", "Prices Pallet Sorting only (1 of 5 cases)", 1, 1, 1),
        (3, "Expansion Consulting Fee", "$75k/partner", "One-time fee per new partner", 1, 0, 0),
        (3, "Expansion Impl Fee/Zone", "$30k", "One-time fee per new zone", 1, 0, 0),
        (3, "NRC First Zone", "$62,000", "CapEx → depreciation; also COGS timing", 0, 1, 1),
        (3, "Monthly Recurring Cost", "$18,000", "COGS driver (recurring per zone)", 0, 1, 1),
        (3, "CapEx Amortization Period", "5 years", "Spreads NRC across years", 0, 1, 0),
        (3, "Maintenance Ramp Start", "40%", "Starting utilization for maintenance", 1, 1, 0),
        (3, "Maintenance Ramp Quarters", "5", "Time to full maintenance utilization", 1, 1, 0),
        (3, "Qtrs After Go-Live (per case)", "0-5", "Delays when each operation type starts", 1, 1, 0),

        # Tier 4 — Low Impact (<$1M EV swing per ±20%)
        (4, "Zone Go-Live Ramp (grid)", "per-partner", "Shifts when zones activate within partner", 1, 0, 0),
        (4, "Planning Start Quarter (per partner)", "Q2'27-Q2'28", "Shifts implementation payment timing", 1, 0, 0),
        (4, "Go-Live Quarter (per partner)", "Q4'27-Q4'28", "Shifts utilization revenue start", 1, 0, 0),
        (4, "License Activation Qtr (per partner)", "Q2'29-Q4'29", "Shifts GAAP vs Non-GAAP cutover", 1, 0, 0),
        (4, "Payment Terms (40/35/25)", "40%/35%/25%", "Redistributes implementation payments", 1, 0, 0),
        (4, "Additional Zone Discount", "25%", "Affects blended cost per zone", 0, 1, 0),
    ]

    TIER_NAMES = {
        1: "Tier 1 — Highest Impact  (>$20M EV swing per ±20% change)",
        2: "Tier 2 — High Impact  ($5–20M EV swing per ±20% change)",
        3: "Tier 3 — Moderate Impact  ($1–5M EV swing per ±20% change)",
        4: "Tier 4 — Low Impact  (<$1M EV swing per ±20% change)",
    }

    COL_HDRS = ["Input", "Default", "Primary Mechanism", "Revenue", "EBITDA", "EV"]
    col_letters = ["B", "C", "D", "E", "F", "G"]

    # Title
    set_cell(ws, 1, "B", "Sensitivity Analysis",
             font=Font(name=FONT_NAME, bold=True, size=14))

    row = 3
    current_tier = 0

    for entry in SENSITIVITY_DATA:
        tier, name, default, mechanism, r, eb, ev = entry

        # New tier header
        if tier != current_tier:
            current_tier = tier
            # Tier header row
            for c in col_letters:
                set_cell(ws, row, c, "", fill=DARK_BLUE_FILL)
            set_cell(ws, row, "B", TIER_NAMES[tier], font=WHITE_BOLD, fill=DARK_BLUE_FILL)
            ws.merge_cells(f"B{row}:G{row}")
            row += 1

            # Column headers
            for ci, hdr in enumerate(COL_HDRS):
                set_cell(ws, row, col_letters[ci], hdr,
                         font=BOLD_FONT, fill=LIGHT_GRAY_FILL)
            row += 1

        # Data row
        set_cell(ws, row, "B", f"  {name}", font=BLACK_FONT)
        set_cell(ws, row, "C", default, font=BLUE_FONT, fill=YELLOW_FILL)
        set_cell(ws, row, "D", mechanism, font=Font(name=FONT_NAME, color="000000", size=9))
        set_cell(ws, row, "E", dot_str(r), font=dot_font(r))
        set_cell(ws, row, "F", dot_str(eb), font=dot_font(eb))
        set_cell(ws, row, "G", dot_str(ev), font=dot_font(ev))
        # Center-align dot columns
        ws[f"E{row}"].alignment = Alignment(horizontal="center")
        ws[f"F{row}"].alignment = Alignment(horizontal="center")
        ws[f"G{row}"].alignment = Alignment(horizontal="center")
        row += 1

    # Notes section
    row += 1
    for c in col_letters:
        set_cell(ws, row, c, "", fill=DARK_BLUE_FILL)
    set_cell(ws, row, "B", "Notes", font=WHITE_BOLD, fill=DARK_BLUE_FILL)
    ws.merge_cells(f"B{row}:G{row}")
    row += 1

    notes = [
        "Impact estimated as ±20% change to input → effect on 5-year output",
        "Revenue = cumulative 5-year GAAP revenue",
        "EBITDA = Year 5 annual EBITDA",
        "EV = EBITDA-based Enterprise Value (Total Company)",
        "",
        "●●● = major impact   ●● = moderate impact   ● = minor impact   — = no direct impact",
    ]
    for note in notes:
        set_cell(ws, row, "B", note, font=Font(name=FONT_NAME, color="808080", size=9))
        row += 1

    return ws


# ============================================================
# MAIN
# ============================================================
def main():
    wb = Workbook()

    build_assumptions(wb)
    build_pipeline_assumptions(wb)
    anchor_ws = build_anchor(wb)
    panchor_ws = build_expansion(wb)
    combined_ws = build_combined(wb, anchor_ws, panchor_ws)
    cost_ws = build_cost_forecast(wb, anchor_ws, panchor_ws)
    build_pnl(wb, anchor_ws, panchor_ws, cost_ws, combined_ws)
    build_valuation_anchor(wb, anchor_ws, combined_ws, cost_ws)
    build_valuation_total(wb, combined_ws, cost_ws)
    build_sources(wb)
    build_sensitivity(wb)

    wb.save(OUTPUT_FILE)
    print(f"Model saved to {OUTPUT_FILE}")
    print("Next: run 'python recalc_win.py arctura_5partner_gtm_model_v1.xlsx' to recalculate formulas")


if __name__ == "__main__":
    main()
